#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from backend.app.vocab.crud.crud_book import book_dao
from backend.app.vocab.model import VocabBookWord, VocabDefinition, VocabExample, VocabWord
from backend.app.vocab.schema.vocab_import import (
    VocabExcelImportResult,
    VocabImportResultItem,
    VocabImportRow,
)
from backend.common.exception import errors
from backend.common.log import log


class VocabImportService:
    """单词 Excel 导入服务"""

    @staticmethod
    async def parse_excel_file(
        *,
        content: bytes,
        filename: str | None,
    ) -> list[VocabImportRow]:
        """
        解析 Excel 文件，返回单词导入行列表

        :param content: 文件二进制内容
        :param filename: 文件名
        :return:
        """
        import io

        import pandas as pd

        from starlette.concurrency import run_in_threadpool

        if not filename or not filename.lower().endswith(('.xlsx', '.xls')):
            raise errors.RequestError(msg='请上传 .xlsx 格式文件')

        excel_bytes = io.BytesIO(content)

        try:
            df = await run_in_threadpool(pd.read_excel, excel_bytes, sheet_name=0)
        except Exception as e:
            raise errors.RequestError(msg=f'读取 Excel 文件失败: {e}')

        df = df.where(df.notna(), None)

        col_map = {
            '单词': '单词',
            '美式音标': '美式音标',
            '英式音标': '英式音标',
            '常用释义': '常用释义',
            '释义1_词性': '释义1_词性',
            '释义1_中文': '释义1_中文',
            '释义2_词性': '释义2_词性',
            '释义2_中文': '释义2_中文',
            '释义3_词性': '释义3_词性',
            '释义3_中文': '释义3_中文',
            '例句1_英文': '例句1_英文',
            '例句1_中文': '例句1_中文',
            '例句2_英文': '例句2_英文',
            '例句2_中文': '例句2_中文',
            '词频等级': '词频等级',
        }

        rows: list[VocabImportRow] = []
        for _, pandas_row in df.iterrows():
            row_dict: dict[str, Any] = {}
            for excel_col, schema_col in col_map.items():
                if excel_col in pandas_row.index:
                    val = pandas_row[excel_col]
                    if pd.notna(val) and val is not None:
                        row_dict[schema_col] = val

            word_text = row_dict.get('单词')
            if not word_text or not str(word_text).strip():
                continue

            row_dict['单词'] = str(word_text).strip()

            # 词频等级转整数
            freq = row_dict.get('词频等级')
            if freq is not None:
                try:
                    row_dict['词频等级'] = int(float(freq))
                except (ValueError, TypeError):
                    row_dict['词频等级'] = 0

            rows.append(VocabImportRow(**row_dict))

        if not rows:
            raise errors.RequestError(msg='Excel 中没有有效单词数据')

        return rows

    @staticmethod
    async def import_from_excel(
        *,
        db: AsyncSession,
        book_id: int | None,
        rows: list[VocabImportRow],
        user_id: int,
    ) -> VocabExcelImportResult:
        """
        从解析后的行数据批量导入单词

        :param db: 数据库会话
        :param book_id: 目标词书 ID，为空则仅创建单词不关联词书
        :param rows: 解析后的导入行
        :param user_id: 操作用户 ID
        :return:
        """
        # 验证词书
        if book_id is not None:
            book = await book_dao.select_model(db, book_id)
            if not book:
                raise errors.NotFoundError(msg='目标词书不存在')
        else:
            book = None

        # 预加载已有单词索引（按 word 文本 -> id）

        existing_stmt = select(VocabWord.id, VocabWord.word).where(VocabWord.word.in_([r.单词 for r in rows]))
        existing_result = await db.execute(existing_stmt)
        existing_map: dict[str, int] = {row.word.lower(): row.id for row in existing_result.all()}

        # 如果要关联词书，预加载已关联的 word_id
        linked_word_ids: set[int] = set()
        if book_id is not None:
            linked_stmt = select(VocabBookWord.word_id).where(VocabBookWord.book_id == book_id)
            linked_result = await db.execute(linked_stmt)
            linked_word_ids = set(linked_result.scalars().all())

        results: list[VocabImportResultItem] = []
        success_count = 0
        created_count = 0
        skipped_count = 0
        linked_count = 0

        for row_index, row in enumerate(rows, start=2):
            try:
                word_lower = row.单词.lower()
                word_id = existing_map.get(word_lower)

                if word_id is not None:
                    # 单词已存在
                    if book_id is not None and word_id not in linked_word_ids:
                        # 已存在单词但未关联到当前词书 → 关联
                        await db.execute(
                            VocabBookWord.__table__.insert().values(
                                book_id=book_id,
                                word_id=word_id,
                                sort_order=row_index,
                            )
                        )
                        linked_word_ids.add(word_id)
                        linked_count += 1
                        results.append(
                            VocabImportResultItem(
                                row_number=row_index,
                                word=row.单词,
                                success=True,
                                action='linked',
                                word_id=word_id,
                            )
                        )
                    else:
                        skipped_count += 1
                        results.append(
                            VocabImportResultItem(
                                row_number=row_index,
                                word=row.单词,
                                success=True,
                                action='skipped',
                                word_id=word_id,
                            )
                        )
                    success_count += 1
                    continue

                # 创建新单词
                word_obj = VocabWord(
                    word=row.单词,
                    phonetic_us=row.美式音标,
                    phonetic_uk=row.英式音标,
                    common_meaning=row.常用释义,
                    frequency=row.词频等级 or 0,
                    created_by=user_id,
                )
                db.add(word_obj)
                await db.flush()
                word_id = word_obj.id

                # 创建释义
                def_order = 0
                for pos, meaning in [
                    (row.释义1_词性, row.释义1_中文),
                    (row.释义2_词性, row.释义2_中文),
                    (row.释义3_词性, row.释义3_中文),
                ]:
                    if not meaning:
                        continue
                    defn = VocabDefinition(
                        word_id=word_id,
                        meaning=str(meaning).strip(),
                        part_of_speech=str(pos).strip() if pos else None,
                        sort_order=def_order,
                    )
                    db.add(defn)
                    def_order += 1

                # 创建例句
                ex_order = 0
                for en, zh in [
                    (row.例句1_英文, row.例句1_中文),
                    (row.例句2_英文, row.例句2_中文),
                ]:
                    if not en:
                        continue
                    ex = VocabExample(
                        word_id=word_id,
                        sentence_en=str(en).strip(),
                        sentence_zh=str(zh).strip() if zh else None,
                        sort_order=ex_order,
                    )
                    db.add(ex)
                    ex_order += 1

                # 关联词书
                if book_id is not None:
                    db.add(
                        VocabBookWord(
                            book_id=book_id,
                            word_id=word_id,
                            sort_order=row_index,
                        )
                    )
                    linked_word_ids.add(word_id)

                existing_map[word_lower] = word_id
                created_count += 1
                success_count += 1
                results.append(
                    VocabImportResultItem(
                        row_number=row_index,
                        word=row.单词,
                        success=True,
                        action='created',
                        word_id=word_id,
                    )
                )

            except Exception as e:
                log.warning(f'单词导入第 {row_index} 行失败: {e}')
                results.append(
                    VocabImportResultItem(
                        row_number=row_index,
                        word=row.单词,
                        success=False,
                        action='error',
                        error_message=str(e),
                    )
                )

        # 更新词书 word_count
        if book is not None:
            new_in_book = created_count + linked_count
            if new_in_book > 0:
                book.word_count = book.word_count + new_in_book

        await db.commit()

        return VocabExcelImportResult(
            total=len(rows),
            success_count=success_count,
            created_count=created_count,
            skipped_count=skipped_count,
            linked_count=linked_count,
            fail_count=len(rows) - success_count,
            details=results,
        )

    @staticmethod
    async def build_import_template() -> bytes:
        """构建 Excel 导入模板"""
        import io

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from starlette.concurrency import run_in_threadpool

        def _build() -> bytes:
            """构建 Excel 模板"""
            wb = Workbook()
            ws = wb.active
            ws.title = '单词导入'

            # 表头
            headers = [
                '单词',
                '美式音标',
                '英式音标',
                '常用释义',
                '释义1_词性',
                '释义1_中文',
                '释义2_词性',
                '释义2_中文',
                '释义3_词性',
                '释义3_中文',
                '例句1_英文',
                '例句1_中文',
                '例句2_英文',
                '例句2_中文',
                '词频等级',
            ]
            ws.append(headers)

            # 表头样式
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)
            for col_idx, _ in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # 示例数据
            ws.append([
                'abandon',
                '/əˈbændən/',
                '/əˈbændən/',
                '放弃；遗弃',
                'v.',
                '放弃，遗弃；沉溺于',
                'n.',
                '放纵，放任',
                None,
                None,
                'He abandoned his wife and children.',
                '他抛弃了妻子和孩子。',
                'The project was abandoned due to lack of funding.',
                '由于缺乏资金，该项目被放弃了。',
                5,
            ])
            ws.append([
                'ability',
                '/əˈbɪləti/',
                '/əˈbɪlɪti/',
                '能力；才能',
                'n.',
                '能力，才能',
                None,
                None,
                None,
                None,
                'She has the ability to solve complex problems.',
                '她有解决复杂问题的能力。',
                None,
                None,
                4,
            ])
            ws.append([
                'abstract',
                '/ˈæbstrækt/',
                '/ˈæbstrækt/',
                '抽象的；摘要',
                'adj.',
                '抽象的，理论性的',
                'n.',
                '摘要，概要',
                'v.',
                '提取，摘录',
                "Abstract art is not to everyone's taste.",
                '抽象艺术并非人人喜欢。',
                None,
                None,
                3,
            ])

            # 列宽
            col_widths = {
                'A': 14,
                'B': 16,
                'C': 16,
                'D': 22,
                'E': 10,
                'F': 24,
                'G': 10,
                'H': 24,
                'I': 10,
                'J': 24,
                'K': 40,
                'L': 30,
                'M': 40,
                'N': 30,
                'O': 10,
            }
            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = width

            # 添加说明 sheet
            ws2 = wb.create_sheet('填写说明')
            instructions = [
                ['字段', '说明', '必填'],
                ['单词', '英文单词原形，不可重复', '是'],
                ['美式音标', '如 /əˈbændən/', '否'],
                ['英式音标', '如 /əˈbændən/', '否'],
                ['常用释义', '简短的常用中文释义（20字以内），用于列表展示', '否'],
                ['释义1_词性', '如 n. / v. / adj. / adv. / prep. / conj.', '否'],
                ['释义1_中文', '该词性下的中文释义，至少填一组释义', '是(至少一组)'],
                ['释义2_词性', '第二个释义的词性', '否'],
                ['释义2_中文', '第二个释义的中文', '否'],
                ['释义3_词性', '第三个释义的词性', '否'],
                ['释义3_中文', '第三个释义的中文', '否'],
                ['例句1_英文', '英文例句', '否'],
                ['例句1_中文', '例句的中文翻译', '否'],
                ['例句2_英文', '第二个英文例句', '否'],
                ['例句2_中文', '第二个例句的中文翻译', '否'],
                ['词频等级', '数字 1-10，数字越大越常见', '否'],
                ['', '', ''],
                ['注意事项', '', ''],
                ['1. 单词列必填，重复单词会自动跳过（仅关联词书）', '', ''],
                ['2. 至少填一组释义（释义1_中文），否则导入后无释义显示', '', ''],
                ['3. 如果单词已存在且选择了词书，会自动关联到词书', '', ''],
                ['4. 支持最多 3 组释义和 2 个例句', '', ''],
            ]
            for row_data in instructions:
                ws2.append(row_data)

            # 说明 sheet 表头样式
            for col_idx in range(1, 4):
                cell = ws2.cell(row=1, column=col_idx)
                cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
                cell.font = Font(color='FFFFFF', bold=True)

            ws2.column_dimensions['A'].width = 14
            ws2.column_dimensions['B'].width = 50
            ws2.column_dimensions['C'].width = 14

            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()

        return await run_in_threadpool(_build)


vocab_import_service: VocabImportService = VocabImportService()
