#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import asyncio
import json
import logging
import re
import uuid
from pathlib import Path
from typing import Any, AsyncGenerator

from sqlalchemy.ext.asyncio import AsyncSession

from backend.app.question_bank.service.parse_service import parse_service
from backend.core.path_conf import UPLOAD_DIR

log = logging.getLogger(__name__)

# 题号正则模式（覆盖常见公考试卷格式）
# 匹配如：1. / 1、 / 1） / （1） / 第1题 / 一、
# 要求：行首开始，且 `数字.` 后面不能紧跟数字（排除小数如 7.5）
QUESTION_NUMBER_PATTERN = re.compile(
    r'(?:^|\n)'                     # 行首
    r'\s*(?:#+\s*)?'                # 可选的前置空格和 Markdown 标题标记（如 # 或 ### ）
    r'(?:'
    r'\d{1,4}\s*[、．）\)]\s*'        # 1、 / 1） 等（非小数风险的标点）
    r'|'
    r'\d{1,4}\s*\.\s*'               # 1. 移除了 (?!\d) 以防止把 4.1916年 漏掉
    r'|'
    r'（\d{1,4}）\s*'                # （1）
    r'|'
    r'第\s*\d{1,4}\s*题[.、：:\s]*'   # 第1题
    r')',
    re.MULTILINE,
)

# AI 提取的系统提示词
SYSTEM_PROMPT = """你是一个专业的公考题库结构化数据提取工具（API）。
你的唯一任务是将用户输入的Markdown格式的公考试卷片段，转换为严格的 JSON 格式。
【最高指令】绝对不要回答试卷中的问题！绝对不要进行评价、解释、分析对错、纠错或寒暄！只做客观的结构化抽取！
这是一份由 Markdown 组成的试卷文本截断（其中的图片链接如 `![图](URL)` 和公式如 `$$...$$` 请完完整整、一字不落的原样保留！严禁改动链接和表格）。

公考中尤其是【资料分析】或【文章阅读】通常会有一大段**公共材料**，接着连续出5道左右选择题。
请提取文本中的所有考题（包括选项、标注的老答案和老解析等），按题号次序严格提取出，并输出为严谨的 JSON。如果没有找到任何题目，必须返回 {"questions": []}。
【注意】如果发现长篇"公共材料（如资料分析文章、表格）"，请**完全忽略并丢弃它们**，本系统的材料将由人工后续录入，你只需要专门提取每一道"单选题"、"多选题"本身！
如果题目附近有明显的所属章节大标题（比如"第一部分 政治理论"、"第二部分 常识判断"等），请一并推断并填入章节名。

**输出 JSON 格式要求规范：**
【严禁编造题型】JSON 中的 `type` 字段必须且只能是以下五种情况之一：
- "single"（单选题）
- "multiple"（多选题）
- "judgement"（判断题）
- "fill"（填空题）
- "shortAnswer"（简答题/申论题/材料分析题/论述题）
如果你解析的为材料大题或论述题，它的 `type` 必须固定写为 "shortAnswer"，绝对不能使用 "material" 或其他非标准短语！

```json
{
  "questions": [
    {
      "type": "single",
      "stem": "<p>...</p>",
      "options_data": {
        "A": {"code": "A", "content": "选项A的内容"},
        "B": {"code": "B", "content": "选项B的内容"}
      },
      "answer_data": {
        "correct": "A"
      },
      "analysis_content": "解析正文...",
      "difficulty": "medium",
      "knowledge_point": "xx考点",
      "score": 1.0,
      "sort_order": 1,
      "source": "卷名来源",
      "year": 2026,
      "chapter_name": "所属章节标题（如果没有则填null）"
    }
  ]
}
```
**千万注意**:
1. 仅仅输出合法且闭合的 JSON 数据对象，必须以 `{` 开始，以 `}` 结束。
2. 不要加任何类似于 ```json ``` 的代码块标记！绝对不要包含任何其它总结、提示性的人类寒暄。
3. 即使你发现原文有逻辑错误或笔误，也不要去修复它，你要做的是原样抓取并JSON化！
"""

# 每批打包的题目数量
BATCH_SIZE = 10
# AI 请求间隔（秒），避免速率限制
AI_REQUEST_INTERVAL = 6


class PipelineService:
    """智能导入流水线服务"""

    # ------------------------------------------------------------------
    #  1. 正则分段引擎
    # ------------------------------------------------------------------

    @staticmethod
    def segment_markdown(md_content: str) -> list[str]:
        """
        按题号正则将 Markdown 切割为独立题目块

        :param md_content: 完整 Markdown 文本
        :return:
        """
        if not md_content or not md_content.strip():
            return []

        # 找到所有题号匹配的位置
        matches = list(QUESTION_NUMBER_PATTERN.finditer(md_content))

        if not matches:
            # 无法识别题号，回退为按段落分割
            log.warning('正则未匹配到题号，回退为段落分割')
            return PipelineService._fallback_segment(md_content)

        segments: list[str] = []
        for i, match in enumerate(matches):
            start = match.start()
            # 跳过匹配前的换行符
            if md_content[start] == '\n':
                start += 1

            end = matches[i + 1].start() if i + 1 < len(matches) else len(md_content)
            # 同样跳过下一个匹配前的换行符
            if end > 0 and md_content[end - 1] == '\n':
                end -= 1

            segment = md_content[start:end].strip()
            if segment:
                segments.append(segment)

        log.info(f'正则分段完成，识别到 {len(segments)} 个题目块')
        return segments

    @staticmethod
    def _fallback_segment(md_content: str) -> list[str]:
        """
        回退分段策略：按双换行分段

        :param md_content: Markdown 文本
        :return:
        """
        paragraphs = re.split(r'\n{2,}', md_content)
        segments = [p.strip() for p in paragraphs if p.strip() and len(p.strip()) > 20]
        log.info(f'回退分段完成，切分为 {len(segments)} 个段落块')
        return segments

    # ------------------------------------------------------------------
    #  2. 批量打包 + AI 提取
    # ------------------------------------------------------------------

    @staticmethod
    async def extract_with_ai(
        db: AsyncSession,
        segments: list[str],
        provider_id: int = 4,
    ) -> AsyncGenerator[dict[str, Any], None]:
        """
        将分段后的题目块批量打包发给 AI 提取，流式 yield 每批的结果和进度

        :param db: 数据库会话
        :param segments: 分段后的题目块列表
        :param provider_id: AI 供应商 ID
        :return:
        """
        from backend.plugin.ai.schema.chat import AIChat, AIChatMessage
        from backend.plugin.ai.service.chat_service import ai_chat_service

        total_segments = len(segments)
        if total_segments == 0:
            yield {'type': 'complete', 'questions': [], 'total_segments': 0}
            return

        # 按 BATCH_SIZE 分组
        batches: list[list[str]] = []
        for i in range(0, total_segments, BATCH_SIZE):
            batches.append(segments[i:i + BATCH_SIZE])

        total_batches = len(batches)
        all_questions: list[dict] = []

        for batch_index, batch in enumerate(batches):
            batch_text = '\n\n---\n\n'.join(batch)
            prompt_text = (
                f'以下是第 {batch_index * BATCH_SIZE + 1} '
                f'到第 {min((batch_index + 1) * BATCH_SIZE, total_segments)} 题的文本片段：\n\n'
                f'{batch_text}\n\n'
                f'请按要求提取所有完整题目。'
            )

            chat_param = AIChat(
                provider_id=provider_id,
                model_id='gpt-5.4',
                messages=[
                    AIChatMessage(role='system', content=SYSTEM_PROMPT),
                    AIChatMessage(role='user', content=prompt_text),
                ],
                temperature=0.1,
            )

            batch_questions: list[dict] = []
            try:
                ai_resp = await ai_chat_service.raw_chat(db=db, chat=chat_param, stream=True)
                resp_content = ai_resp.get('content', '')

                json_match = re.search(r'\{[\s\S]*\}', resp_content)
                if json_match:
                    data = json.loads(json_match.group())
                    batch_questions = data.get('questions', [])
                else:
                    log.warning(f'批次 {batch_index + 1} 未找到合法 JSON: {resp_content[:100]}...')
            except json.JSONDecodeError as je:
                log.error(f'批次 {batch_index + 1} JSON 解析失败: {je}')
            except Exception as e:
                log.error(f'批次 {batch_index + 1} AI 调用失败: {e}')

            all_questions.extend(batch_questions)

            # yield 进度事件
            yield {
                'type': 'progress',
                'batch_index': batch_index + 1,
                'total_batches': total_batches,
                'batch_questions_count': len(batch_questions),
                'total_questions_count': len(all_questions),
            }

            # 限流（最后一批不需要等待）
            if batch_index < total_batches - 1:
                await asyncio.sleep(AI_REQUEST_INTERVAL)

        log.info(f'AI 提取完成，共识别 {len(all_questions)} 道题')

        # yield 最终完成事件
        yield {
            'type': 'complete',
            'questions': all_questions,
            'total_segments': total_segments,
        }

    # ------------------------------------------------------------------
    #  3. 校验规则
    # ------------------------------------------------------------------

    @staticmethod
    def validate_question(q: dict, prev_sort: int | None = None) -> list[str]:
        """
        校验单题数据，返回告警信息列表

        :param q: 题目字典
        :param prev_sort: 上一题的 sort_order
        :return:
        """
        warnings: list[str] = []
        q_type = q.get('type', '')
        stem = q.get('stem', '')
        options = q.get('options_data') or {}
        answer = q.get('answer_data') or {}
        analysis = q.get('analysis_content', '')
        sort_order = q.get('sort_order')

        # 题干为空
        if not stem or not str(stem).strip():
            warnings.append('题干为空')

        # 选项数异常
        if q_type in ('single', 'multiple'):
            valid_options = [k for k, v in options.items() if v]
            if len(valid_options) not in (2, 3, 4, 5):
                warnings.append(f'选项数异常({len(valid_options)}个)')

        # 答案不在选项范围
        correct = answer.get('correct', '')
        if q_type == 'single' and isinstance(correct, str) and correct:
            if correct.upper() not in [k.upper() for k in options.keys()]:
                warnings.append(f'答案{correct}不在选项范围')

        # 判断题选项>2
        if q_type == 'judgement':
            valid_options = [k for k, v in options.items() if v]
            if len(valid_options) > 2:
                warnings.append(f'判断题选项>2({len(valid_options)}个)')

        # 多选题只有 1 答案
        if q_type == 'multiple' and isinstance(correct, list) and len(correct) < 2:
            warnings.append('多选题只有1个答案')

        # 题号跳号
        if prev_sort is not None and sort_order is not None:
            try:
                current = int(sort_order)
                if current != prev_sort + 1:
                    warnings.append(f'题号跳号({prev_sort}→{current})')
            except (ValueError, TypeError):
                pass

        # 解析为空 (允许解析为空，移除校验)
        # if not analysis or not str(analysis).strip():
        #     warnings.append('解析为空')

        return warnings

    # ------------------------------------------------------------------
    #  4. Excel 导出（含校验标红）
    # ------------------------------------------------------------------

    @staticmethod
    def export_to_excel(questions: list[dict], output_path: Path) -> tuple[Path, int]:
        """
        将 AI 提取的 JSON 题目数据导出为 Excel（与 import_from_excel 模板格式一致）

        :param questions: AI 提取的题目列表
        :param output_path: Excel 输出路径
        :return:
        """
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        # 题型反向映射
        type_reverse: dict[str, str] = {
            'single': '单选',
            'multiple': '多选',
            'judgement': '判断',
            'fill': '填空',
            'shortAnswer': '简答',
        }

        # 难度反向映射
        diff_reverse: dict[str, str] = {
            'easy': '简单',
            'medium': '中等',
            'hard': '困难',
        }

        wb = Workbook()
        ws = wb.active
        ws.title = '题目'

        # 表头
        headers = [
            '序号', '题型', '题目', '选项A', '选项B', '选项C', '选项D',
            '答案', '解析', '难度', '分数', '一级目录', '二级目录',
            '知识点', '材料编号', '⚠️校验',
        ]
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # 标红样式
        warn_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        warn_font = Font(color='9C0006')

        warnings_count = 0
        prev_sort: int | None = None

        for row_index, q in enumerate(questions, 2):
            q_type = q.get('type', 'single')
            options = q.get('options_data') or {}
            answer = q.get('answer_data') or {}
            correct = answer.get('correct', '')

            # 将答案格式化为字符串
            if isinstance(correct, list):
                answer_str = ','.join([str(c) for c in correct])
            else:
                answer_str = str(correct) if correct else ''

            # 选项内容提取
            def get_option(code: str) -> str:
                """提取选项内容"""
                opt = options.get(code, {})
                if isinstance(opt, dict):
                    return opt.get('content', '')
                return str(opt) if opt else ''

            sort_order = q.get('sort_order', row_index - 1)

            # 写入数据
            row_data = [
                sort_order,
                type_reverse.get(q_type, q_type),
                q.get('stem', ''),
                get_option('A'),
                get_option('B'),
                get_option('C'),
                get_option('D'),
                answer_str,
                q.get('analysis_content', ''),
                diff_reverse.get(q.get('difficulty', 'medium'), '中等'),
                q.get('score', 1),
                q.get('chapter_name', ''),
                '',  # 二级目录
                '',  # 知识点留白，由用户手动填写
                '',  # 材料编号
            ]

            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_index, column=col, value=value)

            # 校验
            warnings = PipelineService.validate_question(q, prev_sort)
            if warnings:
                warnings_count += 1
                warn_text = '；'.join(warnings)
                ws.cell(row=row_index, column=len(headers), value=warn_text)

                # 标红整行
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_index, column=col)
                    cell.fill = warn_fill
                    cell.font = warn_font

            # 更新 prev_sort
            try:
                prev_sort = int(sort_order)
            except (ValueError, TypeError):
                prev_sort = None

        # 调整列宽
        column_widths = [8, 8, 50, 25, 25, 25, 25, 10, 40, 8, 6, 15, 15, 15, 10, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)].width = width

        # 添加空的“材料” Sheet，供用户手动填写
        ws_material = wb.create_sheet('材料')
        m_headers = ['材料编号', '材料标题', '材料内容']
        for col, header in enumerate(m_headers, 1):
            cell = ws_material.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        ws_material.column_dimensions['A'].width = 15
        ws_material.column_dimensions['B'].width = 30
        ws_material.column_dimensions['C'].width = 80

        # 确保父目录存在
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

        log.info(f'Excel 导出完成: {output_path}，共 {len(questions)} 题，{warnings_count} 条告警')
        return output_path, warnings_count

    # ------------------------------------------------------------------
    #  5. 完整流水线（SSE 流式）
    # ------------------------------------------------------------------

    @staticmethod
    async def run_pipeline(
        *,
        db: AsyncSession,
        file_path: Path,
        file_type: str,
        bank_name: str,
        request_base_url: str,
        provider_id: int = 4,
    ) -> AsyncGenerator[dict[str, Any], None]:
        """
        主流水线：文件 → MD → 分段 → AI 提取 → Excel，全程 SSE 推送进度

        :param db: 数据库会话
        :param file_path: 上传文件路径
        :param file_type: 文件类型（pdf / md）
        :param bank_name: 题库名称（用于文件夹命名）
        :param request_base_url: 请求基础 URL
        :param provider_id: AI 供应商 ID
        :return:
        """
        # ---- 阶段 1：获取 Markdown 并持久化缓存 ----
        yield {'type': 'stage', 'stage': 'parse', 'message': '正在解析文档...'}

        if file_type == 'pdf':
            md_content = await parse_service.parse_pdf_to_markdown(
                file_path=file_path,
                images_dir_name=bank_name,
                request_base_url=request_base_url,
            )
        else:
            md_content = await asyncio.to_thread(file_path.read_text, encoding='utf-8')

        # 保存 Markdown 文件缓存，便于调试核对
        import time
        from pathlib import Path
        export_dir = Path('uploads/pipeline_export')
        export_dir.mkdir(parents=True, exist_ok=True)
        timestamp = time.strftime('%Y%m%d%H%M%S')
        md_file_name = f'{timestamp}_{file_path.stem}.md'
        md_file_path = export_dir / md_file_name
        await asyncio.to_thread(md_file_path.write_text, md_content, encoding='utf-8')

        yield {
            'type': 'stage',
            'stage': 'parse_done',
            'message': f'文档解析完成，共 {len(md_content)} 字符',
            'md_length': len(md_content),
            'md_url': f'/api/v1/qbank/parse/pipeline/download?filename={md_file_name}'
        }

        # ---- 阶段 2：正则分段 ----
        yield {'type': 'stage', 'stage': 'segment', 'message': '正在分段...'}

        segments = PipelineService.segment_markdown(md_content)

        yield {
            'type': 'stage',
            'stage': 'segment_done',
            'message': f'分段完成，识别到 {len(segments)} 个题目块',
            'segments_count': len(segments),
        }

        # ---- 阶段 3：AI 提取 ----
        yield {'type': 'stage', 'stage': 'ai_extract', 'message': '正在进行 AI 智能提取...'}

        all_questions: list[dict] = []
        async for event in PipelineService.extract_with_ai(db, segments, provider_id):
            if event['type'] == 'progress':
                yield event
            elif event['type'] == 'complete':
                all_questions = event.get('questions', [])

        yield {
            'type': 'stage',
            'stage': 'ai_extract_done',
            'message': f'AI 提取完成，共识别 {len(all_questions)} 道题',
            'questions_count': len(all_questions),
        }

        # ---- 阶段 4：生成 Excel ----
        yield {'type': 'stage', 'stage': 'excel', 'message': '正在生成 Excel...'}

        # 去重（基于 sort_order）
        all_questions = PipelineService._deduplicate_questions(all_questions)

        # 生成唯一文件名
        file_id = uuid.uuid4().hex[:8]
        excel_filename = f'pipeline_{bank_name}_{file_id}.xlsx'
        excel_dir = UPLOAD_DIR / 'pipeline_export'
        excel_path = excel_dir / excel_filename

        excel_path, warnings_count = await asyncio.to_thread(
            PipelineService.export_to_excel, all_questions, excel_path,
        )

        # Excel 的下载相对路径
        excel_relative = f'pipeline_export/{excel_filename}'

        yield {
            'type': 'done',
            'message': '流水线执行完成',
            'excel_url': excel_relative,
            'questions_count': len(all_questions),
            'warnings_count': warnings_count,
            'md_length': len(md_content),
            'segments_count': len(segments),
            'questions': all_questions,
        }

    # ------------------------------------------------------------------
    #  6. 去重工具
    # ------------------------------------------------------------------

    @staticmethod
    def _deduplicate_questions(questions: list[dict]) -> list[dict]:
        """
        基于 sort_order + stem 前 50 字符去重

        :param questions: AI 提取的题目列表
        :return:
        """
        seen: set[str] = set()
        unique: list[dict] = []

        for q in questions:
            sort_order = q.get('sort_order', '')
            stem_prefix = str(q.get('stem', ''))[:50].strip()
            key = f'{sort_order}_{stem_prefix}'

            if key in seen:
                continue

            seen.add(key)
            unique.append(q)

        dedup_count = len(questions) - len(unique)
        if dedup_count > 0:
            log.info(f'去重: 原始 {len(questions)} 题 → 去重后 {len(unique)} 题（去除 {dedup_count} 题重复）')

        return unique

    # ------------------------------------------------------------------
    #  7. 分段预览（兜底 API）
    # ------------------------------------------------------------------

    @staticmethod
    def preview_segments(md_content: str) -> list[dict[str, Any]]:
        """
        预览分段结果（不调 AI，仅返回分段 JSON）

        :param md_content: Markdown 文本
        :return:
        """
        segments = PipelineService.segment_markdown(md_content)
        return [
            {
                'index': i,
                'preview': seg[:200] + ('...' if len(seg) > 200 else ''),
                'length': len(seg),
                'content': seg,
            }
            for i, seg in enumerate(segments)
        ]


pipeline_service: PipelineService = PipelineService()
