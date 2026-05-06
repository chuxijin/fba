#!/usr/bin/env python3
import asyncio
import hashlib
import json
import logging
import re

from collections.abc import AsyncGenerator
from pathlib import Path
from typing import Any

from sqlalchemy.ext.asyncio import AsyncSession

from backend.app.question_bank.schema.parse import ReviewExtractResult

log = logging.getLogger(__name__)

# 题号正则模式（覆盖常见公考试卷格式）
# 匹配如：1. / 1、 / 1） / （1） / 第1题 / 一、
# 要求：行首开始，且 `数字.` 后面不能紧跟数字（排除小数如 7.5）
QUESTION_NUMBER_PATTERN = re.compile(
    r'(?:^|\n)'                     # 行首
    r'\s*(?:#+\s*)?'                # 可选的前置空格和 Markdown 标题标记（如 # 或 ### ）
    r'(?:'
    r'\d{1,4}\s*[、．）\)]\s*'        # 1、 / 1） / 1．等
    r'|'
    r'\d{1,4}\s*\.\s*(?!\d)'         # 1. （排除小数如 7.5）
    r'|'
    r'[（(]\d{1,4}[）)]\s*'          # （1） / (1) 全角半角括号
    r'|'
    r'第\s*\d{1,4}\s*题[.、：:\s]*'   # 第1题
    r'|'
    r'[一二三四五六七八九十百]+\s*[、.．]\s*'  # 一、 / 二、 中文数字章节
    r'|'
    r'【\d{1,4}】\s*'                # 【1】方括号题号
    r')',
    re.MULTILINE,
)

# 每批打包的题目数量
BATCH_SIZE = 10
# AI 批次并发数
AI_CONCURRENCY = 5

REVIEW_SYSTEM_PROMPT = """你是一个专业的公考题库结构化抽取工具（API）。
你的任务是把 Markdown 试卷分段转换成严格 JSON，供人工审核台继续校对。
请只返回合法 json 对象，不要输出 Markdown 代码块或额外说明。
绝对不要回答题目，不要纠错，不要补写不存在的内容，只做客观抽取。
图片链接、公式、表格占位必须原样保留。

请同时识别公共材料、完整题目、答案解析条目：
1. 公共材料常见于资料分析、阅读理解、申论材料。材料必须放入 materials。
2. 有题干的完整题目必须放入 questions。如果题目依赖某个材料，请填写 material_id。
3. 每个题目必须带 source_segment_id、question_no_raw、confidence、warnings、status。
4. 每个材料必须带 material_id、title、content、source_segment_ids、confidence、warnings、status。
5. status 默认填 pending_review。
6. confidence 取 0 到 1；不能确定答案或材料绑定时降低置信度并写 warnings。
7. 不要推断 type 和 score；这两个字段会在 Excel 中根据答案自动生成，可以省略。
8. 如果能识别多级篇章，请分别填写 chapter_level1_name、chapter_level2_name、chapter_level3_name；chapter_name 保持为一级篇章用于兼容。
9. 如果分段只有题号、答案和解析，没有题干和选项，绝对不要编造题干，必须放入 answers，不要放入 questions。
10. answers 每条必须带 answer_id、source_segment_id、question_no_raw、sort_order、answer_data、analysis_content、source_quote、confidence、warnings、status。
11. sort_order 必须来自原文题号；如果原文是 37 题，sort_order 就填 37，不能按当前批次内第 1～10 个重排。question_no_raw 保留原文题号文本。

必须输出合法 JSON 对象，不要输出 Markdown 代码块。格式如下：
{
  "materials": [
    {
      "material_id": "M1",
      "title": "材料一",
      "content": "材料正文",
      "source_segment_ids": ["seg_0001"],
      "confidence": 0.85,
      "warnings": [],
      "status": "pending_review"
    }
  ],
  "questions": [
    {
      "question_id": "Q1",
      "source_segment_id": "seg_0002",
      "question_no_raw": "1",
      "stem": "题干",
      "options_data": {
        "A": {"code": "A", "content": "选项A"}
      },
      "answer_data": {"correct": "A"},
      "analysis_content": "解析",
      "difficulty": "medium",
      "knowledge_point": null,
      "sort_order": 1,
      "source": null,
      "year": null,
      "chapter_name": null,
      "chapter_level1_name": null,
      "chapter_level2_name": null,
      "chapter_level3_name": null,
      "material_id": "M1",
      "source_quote": "原文片段",
      "confidence": 0.85,
      "warnings": [],
      "status": "pending_review"
    }
  ],
  "answers": [
    {
      "answer_id": "A1",
      "source_segment_id": "seg_0003",
      "question_no_raw": "1",
      "sort_order": 1,
      "answer_data": {"correct": "A"},
      "analysis_content": "解析正文",
      "source_quote": "原文片段",
      "confidence": 0.85,
      "warnings": ["疑似答案解析册"],
      "status": "pending_review"
    }
  ],
  "warnings": [],
  "needs_review": true
}
"""

ANSWER_REVIEW_SYSTEM_PROMPT = """你是一个专业的答案解析册结构化抽取工具（API）。
你的唯一任务是从 Markdown 中抽取题号、答案和解析，供人工复制粘贴到已有题目。
请只返回合法 json 对象，不要输出 Markdown 代码块或额外说明。
绝对不要把答案解析册识别成新题目；不要编造题干、选项、材料、章节、知识点；不要回答题目。
图片链接、公式、表格占位必须原样保留。

抽取规则：
1. 只输出 answers；materials 和 questions 必须为空数组。
2. 每条 answers 必须带 answer_id、source_segment_id、question_no_raw、sort_order、answer_data、analysis_content、source_quote、confidence、warnings、status。
3. sort_order 必须来自原文题号；如果原文是 37 题，sort_order 就填 37，不能按当前批次内第 1～10 个重排。
4. answer_data.correct 只放答案本身，例如 "A"、["A","C"]、"正确"、"略"。
5. analysis_content 只放解析正文；如果原文答案和解析混在一起，请尽量把答案拆到 answer_data.correct，剩余内容放 analysis_content。
6. 如果只有答案没有解析，analysis_content 填空字符串；如果只有解析没有明确答案，answer_data.correct 填空字符串。
7. status 默认填 pending_review；confidence 取 0 到 1，无法确定题号或答案时降低置信度并写 warnings。

必须输出合法 JSON 对象，不要输出 Markdown 代码块。格式如下：
{
  "materials": [],
  "questions": [],
  "answers": [
    {
      "answer_id": "A1",
      "source_segment_id": "seg_0001",
      "question_no_raw": "1",
      "sort_order": 1,
      "answer_data": {"correct": "A"},
      "analysis_content": "解析正文",
      "source_quote": "原文片段",
      "confidence": 0.85,
      "warnings": [],
      "status": "pending_review"
    }
  ],
  "warnings": [],
  "needs_review": true
}
"""


class ReviewParseService:
    """AI 审核解析服务"""

    # ------------------------------------------------------------------
    #  1. 正则分段引擎
    # ------------------------------------------------------------------

    @staticmethod
    def segment_markdown(md_content: str) -> list[str]:
        """
        按题号正则将 Markdown 切割为独立题目块

        :param md_content: 完整 Markdown 文本
        :return:
        """
        if not md_content or not md_content.strip():
            return []

        # 找到所有题号匹配的位置
        matches = list(QUESTION_NUMBER_PATTERN.finditer(md_content))

        if not matches:
            # 无法识别题号，回退为按段落分割
            log.warning('正则未匹配到题号，回退为段落分割')
            return ReviewParseService._fallback_segment(md_content)

        segments: list[str] = []
        for i, match in enumerate(matches):
            start = match.start()
            # 跳过匹配前的换行符
            if md_content[start] == '\n':
                start += 1

            end = matches[i + 1].start() if i + 1 < len(matches) else len(md_content)
            # 同样跳过下一个匹配前的换行符
            if end > 0 and md_content[end - 1] == '\n':
                end -= 1

            segment = md_content[start:end].strip()
            if segment:
                segments.append(segment)

        log.info(f'正则分段完成，识别到 {len(segments)} 个题目块')
        return segments

    @staticmethod
    def build_review_segments(md_content: str) -> list[dict[str, Any]]:
        """
        构建可追溯审核分段

        :param md_content: Markdown 文本
        :return:
        """
        raw_segments = ReviewParseService.segment_markdown(md_content)
        review_segments: list[dict[str, Any]] = []
        for index, content in enumerate(raw_segments, start=1):
            content_hash = hashlib.sha256(content.encode('utf-8')).hexdigest()
            segment_id = f'seg_{index:04d}'
            review_segments.append({
                'segment_id': segment_id,
                'index': index,
                'type': ReviewParseService._guess_segment_type(content),
                'content': content,
                'preview': content[:200] + ('...' if len(content) > 200 else ''),
                'length': len(content),
                'content_hash': content_hash,
            })
        return review_segments

    @staticmethod
    def _guess_segment_type(content: str) -> str:
        """
        推断分段类型

        :param content: 分段内容
        :return:
        """
        if re.search(r'[AaＡ]\s*[.．、:：]', content) and re.search(r'[BbＢ]\s*[.．、:：]', content):
            return 'question'
        if '根据以下资料' in content or '根据下列资料' in content or '资料' in content[:80]:
            return 'material_or_group'
        return 'unknown'

    @staticmethod
    def _fallback_segment(md_content: str) -> list[str]:
        """
        回退分段策略：按双换行分段

        :param md_content: Markdown 文本
        :return:
        """
        paragraphs = re.split(r'\n{2,}', md_content)
        segments = [p.strip() for p in paragraphs if p.strip() and len(p.strip()) > 20]
        log.info(f'回退分段完成，切分为 {len(segments)} 个段落块')
        return segments

    # ------------------------------------------------------------------
    #  2. 审核分段 + AI 提取
    # ------------------------------------------------------------------

    @staticmethod
    async def extract_review_with_ai(
        db: AsyncSession,
        segments: list[dict[str, Any]],
        provider_id: int = 4,
    ) -> AsyncGenerator[dict[str, Any], None]:
        """
        按审核分段调用 AI，提取材料和题目

        :param db: 数据库会话
        :param segments: 审核分段列表
        :param provider_id: AI 供应商 ID
        :return:
        """
        total_segments = len(segments)
        if total_segments == 0:
            yield {'type': 'complete', 'materials': [], 'questions': [], 'answers': [], 'warnings': []}
            return

        batches: list[list[dict[str, Any]]] = [segments[index:index + BATCH_SIZE] for index in range(0, total_segments, BATCH_SIZE)]

        total_batches = len(batches)
        semaphore = asyncio.Semaphore(AI_CONCURRENCY)

        async def process_batch(batch_index: int, batch: list[dict[str, Any]]) -> dict[str, Any]:
            """
            处理单个审核 AI 批次

            :param batch_index: 批次索引
            :param batch: 审核分段列表
            :return:
            """
            from backend.database.db import async_db_session
            from backend.plugin.ai.schema.chat import AIChat, AIChatMessage
            from backend.plugin.ai.service.chat_service import ai_chat_service

            batch_payload = [
                {
                    'segment_id': item['segment_id'],
                    'index': item['index'],
                    'type': item['type'],
                    'content': item['content'],
                }
                for item in batch
            ]
            prompt_text = (
                f'以下是第 {batch_index + 1}/{total_batches} 批 Markdown 分段，'
                '请只返回合法 json 对象，不要输出 Markdown 或额外文字。'
                '请抽取材料、完整题目、答案解析条目，保留 source_segment_id。'
                '批次内 index 不是原题号；sort_order 必须来自每个分段文本开头的原始题号，'
                'question_no_raw 必须保留原文题号：\n\n'
                f'{json.dumps(batch_payload, ensure_ascii=False)}'
            )

            chat_param = AIChat(
                provider_id=provider_id,
                model_id='gpt-5.4',
                messages=[
                    AIChatMessage(role='system', content=REVIEW_SYSTEM_PROMPT),
                    AIChatMessage(role='user', content=prompt_text),
                ],
                temperature=0.05,
                extra_body={
                    'response_format': {'type': 'json_object'},
                },
            )

            extract_result = ReviewExtractResult()
            batch_warning: str | None = None
            for attempt in range(2):
                try:
                    async with semaphore:
                        async with async_db_session() as batch_db:
                            ai_resp = await ai_chat_service.raw_chat(db=batch_db, chat=chat_param, stream=True)
                            resp_content = ai_resp.get('content', '')
                    json_match = re.search(r'\{[\s\S]*\}', resp_content)
                    if not json_match:
                        raise ValueError('AI 未返回 JSON 对象')
                    data = json.loads(json_match.group())
                    extract_result = ReviewExtractResult.model_validate(data)
                    break
                except Exception as e:
                    batch_warning = f'第 {batch_index + 1} 批第 {attempt + 1} 次抽取失败: {e}'
                    log.warning(batch_warning)
                    if attempt == 0:
                        await asyncio.sleep(2)

            materials = [item.model_dump() for item in extract_result.materials]
            questions = [item.model_dump() for item in extract_result.questions]
            answers = [item.model_dump() for item in extract_result.answers]
            questions, converted_answers = ReviewParseService._split_answer_only_questions(questions)
            answers.extend(converted_answers)
            questions = ReviewParseService._apply_source_question_numbers(
                items=questions,
                source_segments=batch,
            )
            answers = ReviewParseService._apply_source_question_numbers(
                items=answers,
                source_segments=batch,
            )
            materials, questions = ReviewParseService._prefix_review_batch_material_ids(
                batch_index=batch_index,
                materials=materials,
                questions=questions,
                use_prefix=total_batches > 1,
            )
            warnings = list(extract_result.warnings)
            if batch_warning and not extract_result.questions and not extract_result.materials and not extract_result.answers:
                warnings.append(batch_warning)

            return {
                'batch_index': batch_index,
                'materials': materials,
                'questions': questions,
                'answers': answers,
                'warnings': warnings,
            }

        tasks = [
            asyncio.create_task(process_batch(batch_index, batch))
            for batch_index, batch in enumerate(batches)
        ]
        batch_results: list[dict[str, Any] | None] = [None] * total_batches
        completed_batches = 0
        total_materials_count = 0
        total_questions_count = 0
        total_answers_count = 0
        all_warnings: list[str] = []

        for task in asyncio.as_completed(tasks):
            batch_result = await task
            batch_index = int(batch_result['batch_index'])
            materials = batch_result.get('materials', [])
            questions = batch_result.get('questions', [])
            answers = batch_result.get('answers', [])
            warnings = batch_result.get('warnings', [])
            batch_results[batch_index] = batch_result
            completed_batches += 1
            total_materials_count += len(materials)
            total_questions_count += len(questions)
            total_answers_count += len(answers)
            all_warnings.extend(warnings)

            yield {
                'type': 'progress',
                'batch_index': completed_batches,
                'total_batches': total_batches,
                'completed_batch_index': batch_index + 1,
                'batch_materials_count': len(materials),
                'batch_questions_count': len(questions),
                'batch_answers_count': len(answers),
                'total_materials_count': total_materials_count,
                'total_questions_count': total_questions_count,
                'total_answers_count': total_answers_count,
                'warnings': all_warnings,
            }

        all_materials: list[dict[str, Any]] = []
        all_questions: list[dict[str, Any]] = []
        all_answers: list[dict[str, Any]] = []
        for batch_result in batch_results:
            if not batch_result:
                continue
            all_materials.extend(batch_result.get('materials', []))
            all_questions.extend(batch_result.get('questions', []))
            all_answers.extend(batch_result.get('answers', []))

        yield {
            'type': 'complete',
            'materials': ReviewParseService._deduplicate_materials(all_materials),
            'questions': ReviewParseService._deduplicate_questions(all_questions),
            'answers': ReviewParseService._deduplicate_answers(all_answers),
            'warnings': all_warnings,
            'total_segments': total_segments,
        }

    @staticmethod
    async def extract_answers_with_ai(
        db: AsyncSession,
        segments: list[dict[str, Any]],
        provider_id: int = 4,
    ) -> AsyncGenerator[dict[str, Any], None]:
        """
        按答案解析模式调用 AI

        :param db: 数据库会话
        :param segments: 审核分段列表
        :param provider_id: AI 供应商 ID
        :return:
        """
        total_segments = len(segments)
        if total_segments == 0:
            yield {'type': 'complete', 'answers': [], 'warnings': []}
            return

        batches: list[list[dict[str, Any]]] = [segments[index:index + BATCH_SIZE] for index in range(0, total_segments, BATCH_SIZE)]

        total_batches = len(batches)
        semaphore = asyncio.Semaphore(AI_CONCURRENCY)

        async def process_batch(batch_index: int, batch: list[dict[str, Any]]) -> dict[str, Any]:
            """
            处理单个答案解析批次

            :param batch_index: 批次索引
            :param batch: 审核分段列表
            :return:
            """
            from backend.database.db import async_db_session
            from backend.plugin.ai.schema.chat import AIChat, AIChatMessage
            from backend.plugin.ai.service.chat_service import ai_chat_service

            batch_payload = [
                {
                    'segment_id': item['segment_id'],
                    'index': item['index'],
                    'content': item['content'],
                }
                for item in batch
            ]
            prompt_text = (
                f'以下是第 {batch_index + 1}/{total_batches} 批答案解析册 Markdown 分段。'
                '请只返回合法 json 对象，不要输出 Markdown 或额外文字。'
                '批次内 index 不是原题号；sort_order 必须来自每个分段文本开头的原始题号，'
                '只抽取答案和解析，不要抽取题目：\n\n'
                f'{json.dumps(batch_payload, ensure_ascii=False)}'
            )

            chat_param = AIChat(
                provider_id=provider_id,
                model_id='gpt-5.4',
                messages=[
                    AIChatMessage(role='system', content=ANSWER_REVIEW_SYSTEM_PROMPT),
                    AIChatMessage(role='user', content=prompt_text),
                ],
                temperature=0.02,
                extra_body={
                    'response_format': {'type': 'json_object'},
                },
            )

            extract_result = ReviewExtractResult()
            batch_warning: str | None = None
            for attempt in range(2):
                try:
                    async with semaphore:
                        async with async_db_session() as batch_db:
                            ai_resp = await ai_chat_service.raw_chat(db=batch_db, chat=chat_param, stream=True)
                            resp_content = ai_resp.get('content', '')
                    json_match = re.search(r'\{[\s\S]*\}', resp_content)
                    if not json_match:
                        batch_warning = f'第 {batch_index + 1} 批未返回合法 JSON'
                        continue
                    extract_result = ReviewExtractResult.model_validate(json.loads(json_match.group()))
                    break
                except json.JSONDecodeError as e:
                    batch_warning = f'第 {batch_index + 1} 批 JSON 解析失败: {e}'
                except Exception as e:
                    batch_warning = f'第 {batch_index + 1} 批答案解析抽取失败: {e}'
                    if attempt == 0:
                        await asyncio.sleep(2)

            answers = [item.model_dump() for item in extract_result.answers]
            questions = [item.model_dump() for item in extract_result.questions]
            _, converted_answers = ReviewParseService._split_answer_only_questions(questions)
            answers.extend(converted_answers)
            answers = ReviewParseService._apply_source_question_numbers(
                items=answers,
                source_segments=batch,
            )
            warnings = list(extract_result.warnings)
            if batch_warning and not answers:
                warnings.append(batch_warning)

            return {
                'batch_index': batch_index,
                'answers': answers,
                'warnings': warnings,
            }

        tasks = [
            asyncio.create_task(process_batch(batch_index, batch))
            for batch_index, batch in enumerate(batches)
        ]
        batch_results: list[dict[str, Any] | None] = [None] * total_batches
        completed_batches = 0
        total_answers_count = 0
        all_warnings: list[str] = []

        for task in asyncio.as_completed(tasks):
            batch_result = await task
            batch_index = int(batch_result['batch_index'])
            answers = batch_result.get('answers', [])
            warnings = batch_result.get('warnings', [])
            batch_results[batch_index] = batch_result
            completed_batches += 1
            total_answers_count += len(answers)
            all_warnings.extend(warnings)

            yield {
                'type': 'progress',
                'batch_index': completed_batches,
                'total_batches': total_batches,
                'completed_batch_index': batch_index + 1,
                'batch_answers_count': len(answers),
                'total_answers_count': total_answers_count,
                'warnings': all_warnings,
            }

        all_answers: list[dict[str, Any]] = []
        for batch_result in batch_results:
            if not batch_result:
                continue
            all_answers.extend(batch_result.get('answers', []))

        yield {
            'type': 'complete',
            'answers': ReviewParseService._deduplicate_answers(all_answers),
            'warnings': all_warnings,
            'total_segments': total_segments,
        }

    # ------------------------------------------------------------------
    #  3. 校验规则
    # ------------------------------------------------------------------

    @staticmethod
    def validate_question(q: dict, prev_sort: int | None = None) -> list[str]:
        """
        校验单题数据，返回告警信息列表

        :param q: 题目字典
        :param prev_sort: 上一题的 sort_order
        :return:
        """
        warnings: list[str] = []
        q_type = q.get('type', '')
        stem = q.get('stem', '')
        options = q.get('options_data') or {}
        answer = q.get('answer_data') or {}
        q.get('analysis_content', '')
        sort_order = q.get('sort_order')

        # 题干为空
        if not stem or not str(stem).strip():
            warnings.append('题干为空')

        # 选项数异常
        if q_type in ('single', 'multiple'):
            valid_options = [k for k, v in options.items() if v]
            if len(valid_options) not in (2, 3, 4, 5):
                warnings.append(f'选项数异常({len(valid_options)}个)')

        # 答案不在选项范围
        correct = answer.get('correct', '')
        if q_type == 'single' and isinstance(correct, str) and correct:
            if correct.upper() not in [k.upper() for k in options.keys()]:
                warnings.append(f'答案{correct}不在选项范围')

        # 判断题选项>2
        if q_type == 'judgement':
            valid_options = [k for k, v in options.items() if v]
            if len(valid_options) > 2:
                warnings.append(f'判断题选项>2({len(valid_options)}个)')

        # 多选题只有 1 答案
        if q_type == 'multiple' and isinstance(correct, list) and len(correct) < 2:
            warnings.append('多选题只有1个答案')

        # 题号跳号
        if prev_sort is not None and sort_order is not None:
            try:
                current = int(sort_order)
                if current != prev_sort + 1:
                    warnings.append(f'题号跳号({prev_sort}→{current})')
            except (ValueError, TypeError):
                pass

        # 解析为空 (允许解析为空，移除校验)
        # if not analysis or not str(analysis).strip():
        #     warnings.append('解析为空')

        return warnings

    # ------------------------------------------------------------------
    #  4. Excel 导出（含校验标红）
    # ------------------------------------------------------------------

    @staticmethod
    def export_review_to_excel(
        *,
        materials: list[dict[str, Any]],
        questions: list[dict[str, Any]],
        answers: list[dict[str, Any]] | None = None,
        output_path: Path,
    ) -> tuple[Path, int]:
        """
        将审核任务导出为 Excel

        :param materials: 材料列表
        :param questions: 题目列表
        :param answers: 答案解析列表
        :param output_path: 输出路径
        :return:
        """
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        diff_reverse: dict[str, str] = {
            'easy': '简单',
            'medium': '中等',
            'hard': '困难',
        }

        wb = Workbook()
        answer_items = answers or []
        ws = wb.active
        ws.title = '题目'
        headers = [
            '序号', '题型', '题目', '选项A', '选项B', '选项C', '选项D',
            '答案', '解析', '难度', '分数', '一级目录', '二级目录', '三级目录',
            '知识点', '材料编号', '置信度', '审核状态', '⚠️校验',
        ]
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        warn_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        warn_font = Font(color='9C0006')

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        warnings_count = 0
        prev_sort: int | None = None
        for row_index, question in enumerate(questions, 2):
            options = question.get('options_data') or {}
            answer = question.get('answer_data') or {}
            level1_name, level2_name, level3_name = ReviewParseService._get_question_chapter_names(question)
            correct = answer.get('correct', '')
            if isinstance(correct, list):
                answer_str = ','.join([str(item) for item in correct])
            else:
                answer_str = str(correct) if correct else ''

            def get_option(code: str) -> str:
                """提取选项内容"""
                option = options.get(code, {})
                if isinstance(option, dict):
                    return option.get('content', '')
                return str(option) if option else ''

            row_data = [
                question.get('sort_order', row_index - 1),
                f'=IF(ISBLANK(H{row_index}), "", IF(LEN(TRIM(H{row_index}))>1, "多选", "单选"))',
                question.get('stem', ''),
                get_option('A'),
                get_option('B'),
                get_option('C'),
                get_option('D'),
                answer_str,
                question.get('analysis_content', ''),
                diff_reverse.get(question.get('difficulty', 'medium'), '中等'),
                f'=IF(B{row_index}="多选", 2, IF(B{row_index}="单选", 1, ""))',
                level1_name,
                level2_name,
                level3_name,
                question.get('knowledge_point', ''),
                question.get('material_id', ''),
                question.get('confidence', ''),
                question.get('status', ''),
            ]
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_index, column=col, value=value)

            warnings = ReviewParseService.validate_question(question, prev_sort)
            warnings.extend(question.get('warnings') or [])
            if warnings:
                warnings_count += 1
                ws.cell(row=row_index, column=len(headers), value='；'.join(warnings))
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_index, column=col)
                    cell.fill = warn_fill
                    cell.font = warn_font

            try:
                prev_sort = int(question.get('sort_order'))
            except (ValueError, TypeError):
                prev_sort = None

        column_widths = [8, 8, 50, 25, 25, 25, 25, 10, 40, 8, 6, 15, 15, 15, 15, 12, 10, 12, 30]
        for index, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + index) if index <= 26 else 'A' + chr(64 + index - 26)].width = width

        ws_material = wb.create_sheet('材料')
        material_headers = ['材料编号', '材料标题', '材料内容', '置信度', '审核状态', '来源分段', '⚠️校验']
        for col, header in enumerate(material_headers, 1):
            cell = ws_material.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for row_index, material in enumerate(materials, 2):
            source_ids = material.get('source_segment_ids') or []
            material_warnings = material.get('warnings') or []
            ws_material.append([
                material.get('material_id', ''),
                material.get('title', ''),
                material.get('content', ''),
                material.get('confidence', ''),
                material.get('status', ''),
                ','.join([str(item) for item in source_ids]),
                '；'.join(material_warnings),
            ])
            if material_warnings:
                warnings_count += 1
                for col in range(1, len(material_headers) + 1):
                    cell = ws_material.cell(row=row_index, column=col)
                    cell.fill = warn_fill
                    cell.font = warn_font

        ws_material.column_dimensions['A'].width = 15
        ws_material.column_dimensions['B'].width = 30
        ws_material.column_dimensions['C'].width = 80
        ws_material.column_dimensions['D'].width = 10
        ws_material.column_dimensions['E'].width = 12
        ws_material.column_dimensions['F'].width = 30
        ws_material.column_dimensions['G'].width = 30

        ws_answer = wb.create_sheet('答案解析')
        answer_headers = ['解析编号', '题号', '答案', '解析', '置信度', '审核状态', '来源分段', '原文片段', '⚠️校验']
        for col, header in enumerate(answer_headers, 1):
            cell = ws_answer.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for row_index, answer_item in enumerate(answer_items, 2):
            answer_data = answer_item.get('answer_data') or {}
            correct = answer_data.get('correct', '')
            if isinstance(correct, list):
                answer_str = ','.join([str(item) for item in correct])
            else:
                answer_str = str(correct) if correct else ''
            answer_warnings = answer_item.get('warnings') or []
            ws_answer.append([
                answer_item.get('answer_id', ''),
                answer_item.get('question_no_raw') or answer_item.get('sort_order') or '',
                answer_str,
                answer_item.get('analysis_content', ''),
                answer_item.get('confidence', ''),
                answer_item.get('status', ''),
                answer_item.get('source_segment_id', ''),
                answer_item.get('source_quote', ''),
                '；'.join(answer_warnings),
            ])
            if answer_warnings:
                warnings_count += 1
                for col in range(1, len(answer_headers) + 1):
                    cell = ws_answer.cell(row=row_index, column=col)
                    cell.fill = warn_fill
                    cell.font = warn_font

        ws_answer.column_dimensions['A'].width = 14
        ws_answer.column_dimensions['B'].width = 12
        ws_answer.column_dimensions['C'].width = 12
        ws_answer.column_dimensions['D'].width = 80
        ws_answer.column_dimensions['E'].width = 10
        ws_answer.column_dimensions['F'].width = 12
        ws_answer.column_dimensions['G'].width = 18
        ws_answer.column_dimensions['H'].width = 60
        ws_answer.column_dimensions['I'].width = 30

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        return output_path, warnings_count

    # ------------------------------------------------------------------
    #  5. 去重工具
    # ------------------------------------------------------------------

    @staticmethod
    def _extract_question_number_from_segment(content: str) -> tuple[str, int] | None:
        """
        从原始分段开头提取题号

        :param content: 分段文本
        :return:
        """
        match = re.match(
            r'^\s*(?:#+\s*)?(?:'
            r'第\s*(\d{1,4})\s*题'
            r'|[（(](\d{1,4})[）)]'
            r'|【(\d{1,4})】'
            r'|(\d{1,4})\s*[、．）\)]'
            r'|(\d{1,4})\s*\.\s*(?!\d)'
            r')',
            content,
        )
        if not match:
            return None

        number_text = next((item for item in match.groups() if item), None)
        if not number_text:
            return None
        return number_text, int(number_text)

    @staticmethod
    def _apply_source_question_numbers(
        *,
        items: list[dict[str, Any]],
        source_segments: list[str | dict[str, Any]],
    ) -> list[dict[str, Any]]:
        """
        用原始分段题号修正 AI 批次内重排

        :param items: AI 抽取条目
        :param source_segments: 原始分段
        :return:
        """
        number_by_segment_id: dict[str, tuple[str, int]] = {}
        numbers_by_index: list[tuple[str, int] | None] = []
        for segment in source_segments:
            if isinstance(segment, dict):
                content = str(segment.get('content') or '')
                segment_id = str(segment.get('segment_id') or '')
            else:
                content = str(segment)
                segment_id = ''

            number_info = ReviewParseService._extract_question_number_from_segment(content)
            numbers_by_index.append(number_info)
            if segment_id and number_info:
                number_by_segment_id[segment_id] = number_info

        for index, item in enumerate(items):
            source_segment_id = str(item.get('source_segment_id') or '')
            number_info = number_by_segment_id.get(source_segment_id)
            if not number_info and index < len(numbers_by_index):
                number_info = numbers_by_index[index]
            if not number_info:
                continue

            question_no_raw, sort_order = number_info
            item['question_no_raw'] = question_no_raw
            item['sort_order'] = sort_order
        return items

    @staticmethod
    def _deduplicate_questions(questions: list[dict]) -> list[dict]:
        """
        基于 sort_order + stem 前 50 字符去重

        :param questions: AI 提取的题目列表
        :return:
        """
        seen: set[str] = set()
        unique: list[dict] = []

        for q in questions:
            sort_order = q.get('sort_order', '')
            stem_prefix = str(q.get('stem', ''))[:50].strip()
            key = f'{sort_order}_{stem_prefix}'

            if key in seen:
                continue

            seen.add(key)
            unique.append(q)

        dedup_count = len(questions) - len(unique)
        if dedup_count > 0:
            log.info(f'去重: 原始 {len(questions)} 题 → 去重后 {len(unique)} 题（去除 {dedup_count} 题重复）')

        return unique

    @staticmethod
    def _deduplicate_answers(answers: list[dict[str, Any]]) -> list[dict[str, Any]]:
        """
        基于题号和解析内容去重

        :param answers: 答案解析列表
        :return:
        """
        seen: set[str] = set()
        unique: list[dict[str, Any]] = []
        for answer_item in answers:
            question_no = str(answer_item.get('question_no_raw') or answer_item.get('sort_order') or '').strip()
            analysis_prefix = str(answer_item.get('analysis_content') or '')[:80].strip()
            correct = answer_item.get('answer_data') or {}
            answer_value = str(correct.get('correct') or '').strip() if isinstance(correct, dict) else ''
            key = f'{question_no}_{answer_value}_{analysis_prefix}'
            if not key.strip('_') or key in seen:
                continue
            seen.add(key)
            unique.append(answer_item)
        return unique

    @staticmethod
    def _split_answer_only_questions(
        questions: list[dict[str, Any]],
    ) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
        """
        将只有答案解析的伪题目拆到答案解析列表

        :param questions: 题目列表
        :return:
        """
        complete_questions: list[dict[str, Any]] = []
        answer_items: list[dict[str, Any]] = []
        for index, question in enumerate(questions, start=1):
            stem = str(question.get('stem') or '').strip()
            options = question.get('options_data') or {}
            answer_data = question.get('answer_data') or {}
            analysis_content = str(question.get('analysis_content') or '').strip()
            if stem or options:
                complete_questions.append(question)
                continue
            if not answer_data and not analysis_content:
                complete_questions.append(question)
                continue

            warnings = question.get('warnings') or []
            warning = '疑似答案解析册'
            if warning not in warnings:
                warnings.append(warning)
            answer_items.append({
                'answer_id': str(question.get('question_id') or f'A{index}').replace('Q', 'A', 1),
                'source_segment_id': question.get('source_segment_id'),
                'question_no_raw': question.get('question_no_raw'),
                'sort_order': question.get('sort_order') or index,
                'answer_data': answer_data,
                'analysis_content': analysis_content,
                'source_quote': question.get('source_quote'),
                'confidence': question.get('confidence', 0.5),
                'warnings': warnings,
                'status': question.get('status') or 'pending_review',
            })
        return complete_questions, answer_items

    @staticmethod
    def _deduplicate_materials(materials: list[dict[str, Any]]) -> list[dict[str, Any]]:
        """
        基于 material_id 和内容去重

        :param materials: 材料列表
        :return:
        """
        seen: set[str] = set()
        unique: list[dict[str, Any]] = []
        for material in materials:
            material_id = str(material.get('material_id') or '')
            content_prefix = str(material.get('content') or '')[:80].strip()
            key = material_id or content_prefix
            if not key or key in seen:
                continue
            seen.add(key)
            unique.append(material)
        return unique

    @staticmethod
    def _prefix_review_batch_material_ids(
        *,
        batch_index: int,
        materials: list[dict[str, Any]],
        questions: list[dict[str, Any]],
        use_prefix: bool,
    ) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
        """
        为审核批次材料编号加前缀

        :param batch_index: 批次索引
        :param materials: 材料列表
        :param questions: 题目列表
        :param use_prefix: 是否加前缀
        :return:
        """
        if not use_prefix:
            return materials, questions

        material_id_map: dict[str, str] = {}
        prefix = f'B{batch_index + 1:03d}'
        for index, material in enumerate(materials, start=1):
            raw_material_id = str(material.get('material_id') or f'M{index}')
            new_material_id = f'{prefix}_{raw_material_id}'
            material_id_map[raw_material_id] = new_material_id
            material['material_id'] = new_material_id

        for question in questions:
            raw_material_id = question.get('material_id')
            if raw_material_id is None:
                continue
            question['material_id'] = material_id_map.get(str(raw_material_id), raw_material_id)

        return materials, questions

    @staticmethod
    def _get_question_chapter_names(question: dict[str, Any]) -> tuple[str, str, str]:
        """
        获取题目多级章节名

        :param question: 题目数据
        :return:
        """
        level1_name = (
            question.get('chapter_level1_name')
            or question.get('一级目录')
            or question.get('chapter_name')
            or ''
        )
        level2_name = question.get('chapter_level2_name') or question.get('二级目录') or ''
        level3_name = question.get('chapter_level3_name') or question.get('三级目录') or ''
        return str(level1_name), str(level2_name), str(level3_name)


review_parse_service: ReviewParseService = ReviewParseService()
