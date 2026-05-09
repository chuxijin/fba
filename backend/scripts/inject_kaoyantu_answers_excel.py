#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将 Burp 答案解析 XML 注入已有题目 Excel

用法:
    python backend/scripts/inject_kaoyantu_answers_excel.py --answers C:\\Users\\19396\\Desktop\\answer
    python backend/scripts/inject_kaoyantu_answers_excel.py --answers answer.xml --input questions.xlsx --output questions_with_answers.xlsx
"""
import argparse
import base64
import json
import xml.etree.ElementTree as ET
from collections import OrderedDict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment

DEFAULT_INPUT = 'backend/scripts/outputs/kaoyantu_questions_738_all_pages.xlsx'
DEFAULT_OUTPUT = 'backend/scripts/outputs/kaoyantu_questions_738_all_pages_with_answers.xlsx'


def decode_burp_node_text(node: ET.Element) -> str:
    """
    解码 Burp 节点文本

    :param node: Burp request 或 response 节点
    :return: 不添加返回说明
    """
    text = node.text or ''
    if node.attrib.get('base64') == 'true':
        return base64.b64decode(text).decode('utf-8', errors='ignore')

    return text


def extract_http_body(raw_http: str) -> str:
    """
    提取 HTTP 正文

    :param raw_http: 原始 HTTP 文本
    :return: 不添加返回说明
    """
    for separator in ('\r\n\r\n', '\n\n'):
        if separator in raw_http:
            return raw_http.split(separator, 1)[1].strip()

    return raw_http.strip()


def parse_json_body(raw_http: str) -> dict[str, Any] | None:
    """
    解析 HTTP 正文 JSON

    :param raw_http: 原始 HTTP 文本
    :return: 不添加返回说明
    """
    body = extract_http_body(raw_http)
    if not body:
        return None

    try:
        parsed = json.loads(body)
    except json.JSONDecodeError:
        return None

    if isinstance(parsed, dict):
        return parsed

    return None


def collect_answer_comments(answer_xml_path: Path) -> OrderedDict[int, str]:
    """
    收集答案解析

    :param answer_xml_path: Burp XML 文件路径
    :return: 不添加返回说明
    """
    tree = ET.parse(answer_xml_path)
    comments_map: OrderedDict[int, str] = OrderedDict()

    for item in tree.findall('.//item'):
        response_node = item.find('response')
        if response_node is None:
            continue

        response_json = parse_json_body(decode_burp_node_text(response_node))
        if not response_json:
            continue

        data = response_json.get('data')
        if not isinstance(data, list):
            continue

        for answer_item in data:
            if not isinstance(answer_item, dict):
                continue

            question_id = answer_item.get('id')
            comments = answer_item.get('comments')
            if not isinstance(question_id, int) or not comments:
                continue

            comments_map[question_id] = str(comments)

    return comments_map


def get_header_index_map(ws: Any) -> dict[str, int]:
    """
    获取表头列索引

    :param ws: 工作表对象
    :return: 不添加返回说明
    """
    header_map: dict[str, int] = {}
    for cell in ws[1]:
        if cell.value:
            header_map[str(cell.value)] = cell.column

    return header_map


def inject_by_question_id(
    ws: Any,
    comments_map: OrderedDict[int, str],
    id_column: int,
    analysis_column: int,
) -> int:
    """
    按题目 ID 注入解析

    :param ws: 工作表对象
    :param comments_map: 题目 ID 与解析映射
    :param id_column: ID 列
    :param analysis_column: 解析列
    :return: 不添加返回说明
    """
    injected_count = 0
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    for row_number in range(2, ws.max_row + 1):
        raw_id = ws.cell(row=row_number, column=id_column).value
        try:
            question_id = int(raw_id)
        except (TypeError, ValueError):
            continue

        comments = comments_map.get(question_id)
        if not comments:
            continue

        cell = ws.cell(row=row_number, column=analysis_column)
        cell.value = comments
        cell.alignment = wrap_alignment
        injected_count += 1

    return injected_count


def inject_by_order(ws: Any, comments_map: OrderedDict[int, str], analysis_column: int) -> int:
    """
    按答案出现顺序注入解析

    :param ws: 工作表对象
    :param comments_map: 题目 ID 与解析映射
    :param analysis_column: 解析列
    :return: 不添加返回说明
    """
    injected_count = 0
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    comments_values = list(comments_map.values())
    for index, comments in enumerate(comments_values, start=2):
        if index > ws.max_row:
            break

        title = ws.cell(row=index, column=3).value
        if not title:
            continue

        cell = ws.cell(row=index, column=analysis_column)
        cell.value = comments
        cell.alignment = wrap_alignment
        injected_count += 1

    return injected_count


def inject_answers_to_excel(
    *,
    answer_xml_path: Path,
    input_path: Path,
    output_path: Path,
    force_order: bool,
) -> dict[str, int | str]:
    """
    注入答案解析到 Excel

    :param answer_xml_path: 答案 XML 路径
    :param input_path: 输入 Excel 路径
    :param output_path: 输出 Excel 路径
    :param force_order: 是否强制按顺序注入
    :return: 不添加返回说明
    """
    comments_map = collect_answer_comments(answer_xml_path)
    if not comments_map:
        raise RuntimeError('未从答案 XML 中提取到 comments 解析')

    workbook = load_workbook(input_path)
    ws = workbook['题目']
    header_map = get_header_index_map(ws)
    analysis_column = header_map.get('解析')
    if not analysis_column:
        raise RuntimeError('Excel 中未找到 解析 列')

    injected_count = 0
    mode = 'order'
    id_column = header_map.get('题目ID') or header_map.get('id') or header_map.get('ID') or header_map.get('序号')
    if id_column and not force_order:
        injected_count = inject_by_question_id(
            ws=ws,
            comments_map=comments_map,
            id_column=id_column,
            analysis_column=analysis_column,
        )
        mode = 'id'

    if injected_count == 0:
        injected_count = inject_by_order(
            ws=ws,
            comments_map=comments_map,
            analysis_column=analysis_column,
        )
        mode = 'order'

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    return {
        'answers': len(comments_map),
        'injected': injected_count,
        'mode': mode,
        'input': str(input_path),
        'output': str(output_path),
    }


def main() -> None:
    """执行解析注入"""
    parser = argparse.ArgumentParser(description='将 Burp 答案解析 XML 注入已有题目 Excel')
    parser.add_argument('--answers', required=True, help='Burp 答案 XML 文件路径')
    parser.add_argument('--input', default=DEFAULT_INPUT, help='已有题目 Excel 路径')
    parser.add_argument('--output', default=DEFAULT_OUTPUT, help='输出 Excel 路径')
    parser.add_argument('--force-order', action='store_true', help='强制按答案出现顺序写入解析')
    args = parser.parse_args()

    result = inject_answers_to_excel(
        answer_xml_path=Path(args.answers),
        input_path=Path(args.input),
        output_path=Path(args.output),
        force_order=args.force_order,
    )
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()
