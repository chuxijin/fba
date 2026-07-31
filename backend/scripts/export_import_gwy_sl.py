#!/usr/bin/env python3
"""导出国考申论真题（2013-2026）到 Excel，导入本地 V2，并复制到业务目录。"""

import argparse
import json
import os
import random
import sys
import time
from decimal import Decimal
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
os.environ['FBA_DEV'] = '1'

from sqlalchemy import create_engine, text
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

COLUMNS = [
    'question_type', 'stem', 'answer', 'explanation_default', 'explanation_official', 'explanation_expert',
    'option_A', 'option_B', 'option_C', 'option_D', 'option_E',
    'score',
    'section_l1', 'section_l2', 'section_l3',
    'knowledge_point', 'item_key',
]

V1_TO_V2_TYPE = {
    'single': 'single_choice', 'multiple': 'multiple_choice',
    'judgement': 'true_false', 'fill': 'fill_blank', 'shortAnswer': 'short_answer',
}


def safe_str(val):
    if val is None:
        return ''
    s = str(val)
    s = ''.join(c if c == '\t' or c == '\n' or c == '\r' or (ord(c) >= 32) else ' ' for c in s)
    if len(s) > 32767:
        s = s[:32767]
    return s


def sanitize_filename(name):
    for ch in ['/', '\\', ':', '*', '?', '"', '<', '>', '|']:
        name = name.replace(ch, '_')
    return name.strip()


def get_engine(db_url=None):
    if db_url:
        return create_engine(db_url, pool_pre_ping=True)
    from backend.core.conf import settings
    driver = 'psycopg' if settings.DATABASE_TYPE == 'postgresql' else 'pymysql'
    database = settings.DATABASE_SCHEMA
    url = (
        f"{settings.DATABASE_TYPE}+{driver}://{settings.DATABASE_USER}:{settings.DATABASE_PASSWORD}"
        f"@{settings.DATABASE_HOST}:{settings.DATABASE_PORT}/{database}"
    )
    return create_engine(url, pool_pre_ping=True)


def fetch_banks(conn):
    rows = conn.execute(text(
        "SELECT id, code, name, year "
        "FROM fba.study_question_bank "
        "WHERE code LIKE 'PAPER_GWY_SL_NAT_%' AND deleted = 0 AND status = 1 "
        "ORDER BY year DESC, id"
    )).mappings().all()
    return [dict(r) for r in rows]


def fetch_placements(conn, bank_id):
    rows = conn.execute(text(
        "SELECT p.question_id, p.chapter_id, p.sort_order "
        "FROM fba.study_question_placement p "
        "WHERE p.bank_id = :bid AND p.deleted = 0 AND p.is_active = true "
        "ORDER BY p.chapter_id, p.sort_order"
    ), {'bid': bank_id}).mappings().all()
    return [dict(r) for r in rows]


def fetch_questions(conn, qids):
    result = {}
    if not qids:
        return result
    for i in range(0, len(qids), 2000):
        batch = qids[i:i + 2000]
        rows = conn.execute(text(
            "SELECT q.id, q.type, q.stem, q.options, q.default_score, q.knowledge_point "
            "FROM fba.study_question q WHERE q.id = ANY(:qids) AND q.deleted = 0"
        ), {'qids': batch}).mappings().all()
        for r in rows:
            result[r['id']] = dict(r)
    return result


def fetch_analyses(conn, qids):
    result = {}
    if not qids:
        return result
    for i in range(0, len(qids), 2000):
        batch = qids[i:i + 2000]
        rows = conn.execute(text(
            "SELECT a.question_id, a.answer_data, a.content "
            "FROM fba.study_question_analysis a "
            "WHERE a.question_id = ANY(:qids) AND a.deleted = 0 AND a.is_default = true"
        ), {'qids': batch}).mappings().all()
        for r in rows:
            result[r['question_id']] = dict(r)
    return result


def parse_options(options_json):
    if not options_json:
        return {}
    opts = json.loads(options_json) if isinstance(options_json, str) else options_json
    result = {}
    if isinstance(opts, dict):
        for k, v in opts.items():
            label = k.strip().upper()
            if label and v and label in 'ABCDE':
                result[f'option_{label}'] = safe_str(v)
    return result


def parse_answer(answer_data):
    if not answer_data:
        return ''
    data = json.loads(answer_data) if isinstance(answer_data, str) else answer_data
    if isinstance(data, dict):
        answers = data.get('answer') or data.get('answers') or data.get('correct') or ''
        if isinstance(answers, list):
            return safe_str('\n'.join(str(a) for a in answers))
        return safe_str(answers)
    if isinstance(data, list):
        return safe_str('\n'.join(str(a) for a in data))
    return safe_str(data)


def export_to_excel(engine, banks, output_dir):
    files = []
    with engine.connect() as conn:
        for bank in banks:
            placements = fetch_placements(conn, bank['id'])
            if not placements:
                print(f'  [{bank["name"]}] no questions')
                continue
            qids = list(set(p['question_id'] for p in placements))
            questions = fetch_questions(conn, qids)
            analyses = fetch_analyses(conn, qids)

            wb = Workbook()
            ws = wb.active
            ws.title = '题目'
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            thin = Border(*[Side(style='thin')] * 4)
            for ci, cn in enumerate(COLUMNS, 1):
                cell = ws.cell(row=1, column=ci, value=cn)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin

            row_num = 2
            for p in placements:
                q = questions.get(p['question_id'])
                if not q:
                    continue
                analysis = analyses.get(p['question_id'], {})
                row_data = {
                    'question_type': V1_TO_V2_TYPE.get(q['type'], q['type']),
                    'stem': safe_str(q['stem'] or ''),
                    'answer': parse_answer(analysis.get('answer_data')),
                    'explanation_default': safe_str(analysis.get('content', '')),
                    'option_A': '', 'option_B': '', 'option_C': '', 'option_D': '', 'option_E': '',
                    'score': safe_str(q.get('default_score', '1')),
                    'section_l1': '申论',
                    'section_l2': '',
                    'section_l3': '',
                    'knowledge_point': '',
                    'item_key': f'sl_{bank["year"]}_{bank["id"]}_{p["question_id"]}',
                }
                for ci, cn in enumerate(COLUMNS, 1):
                    val = row_data.get(cn, '')
                    cell = ws.cell(row=row_num, column=ci, value=val)
                    cell.border = thin
                    if cn in ('stem', 'answer', 'explanation_default'):
                        cell.alignment = Alignment(wrap_text=True)
                row_num += 1

            fname = f'{bank["year"]}_{bank["code"]}_{sanitize_filename(bank["name"])}.xlsx'
            fpath = output_dir / fname
            wb.save(fpath)
            files.append(fpath)
            print(f'  Exported {row_num - 2} questions → {fname}')
    return files


def parse_explanations(row):
    result = []
    for key, val in row.items():
        k = key.strip().replace(' ', '_').replace('-', '_')
        if k.startswith('explanation_'):
            exp_type = k[len('explanation_'):]
            if exp_type and val and str(val).strip():
                result.append({'explanation_type': exp_type, 'content': str(val).strip()})
    if not result:
        result.append({'explanation_type': 'default', 'content': '暂无解析'})
    return result


def parse_answer_v2(answer_str, qtype):
    if not answer_str:
        return {'correct': ''}
    answer_str = answer_str.strip()
    if qtype == 'single_choice':
        import re
        codes = re.findall(r'[A-Za-z]', answer_str.upper())
        return {'correct': codes[0]} if codes else {'correct': answer_str}
    if qtype == 'multiple_choice':
        import re
        codes = re.findall(r'[A-Za-z]', answer_str.upper())
        return {'correct': sorted(set(codes))} if codes else {'correct': [answer_str]}
    if qtype == 'true_false':
        return {'correct': answer_str.lower() in ('对', '正确', 'true', 't', '1', '是')}
    parts = [p.strip() for p in answer_str.replace('，', ',').split('\n') if p.strip()]
    return {'correct': parts if parts else [answer_str]}


def parse_options_v2(row):
    opts = []
    for letter in 'ABCDE':
        val = row.get(f'option_{letter}')
        if val and str(val).strip():
            opts.append({'option_code': letter, 'content': str(val).strip(), 'sort_order': ord(letter) - 65})
    return opts


def fix_sequences(conn):
    tables = [
        'qbank_v2_bank', 'qbank_v2_bank_revision', 'qbank_v2_bank_item',
        'qbank_v2_bank_section', 'qbank_v2_question', 'qbank_v2_question_answer',
        'qbank_v2_question_explanation',
    ]
    for table in tables:
        conn.execute(text(f"SELECT setval('{table}_id_seq', COALESCE((SELECT MAX(id) FROM {table}), 0) + 1)"))
    conn.commit()


def import_file(conn, file_path):
    import pandas as pd
    df = pd.read_excel(file_path, sheet_name=0, dtype=str)
    df = df.where(df.notna(), None)

    stem = Path(file_path).stem
    parts = stem.split('_', 3)
    code_base = parts[1] if len(parts) > 1 else 'imp'
    display_name = stem
    import re
    m = re.match(r'\d{4}_[A-Z_]+_(.+)', stem)
    if m:
        display_name = m.group(1)

    rows = []
    for _, r in df.iterrows():
        row = {}
        for col in df.columns:
            cc = str(col).strip().replace(' ', '_').replace('-', '_')
            val = r[col]
            if pd.notna(val) and val is not None:
                row[cc] = str(val).strip()
        if not row.get('stem'):
            continue
        rows.append(row)

    if not rows:
        print(f'    No valid rows')
        return

    def item_key_exists(key):
        if not key:
            return False
        return conn.execute(text("SELECT 1 FROM qbank_v2_bank_item WHERE item_key = :k AND deleted = 0 LIMIT 1"), {'k': key}).fetchone() is not None

    if all(item_key_exists(r.get('item_key', '')) for r in rows if r.get('item_key')):
        print(f'    All {len(rows)} questions already exist, skipped')
        return

    ts = int(time.time())
    code = f'imp_sl_{ts}_{random.randint(1000, 9999)}'
    bank_id = conn.execute(text(
        "INSERT INTO qbank_v2_bank (code, owner_id, visibility, status, created_by, created_time) "
        "VALUES (:code, NULL, 'public', 'active', 1, NOW()) RETURNING id"
    ), {'code': code}).fetchone()[0]

    revision_id = conn.execute(text(
        "INSERT INTO qbank_v2_bank_revision "
        "(bank_id, revision_no, name, bank_kind, settings, question_count, total_score, status, created_by, created_time) "
        "VALUES (:bid, 1, :name, 'paper', '{}'::jsonb, 0, 0, 'draft', 1, NOW()) RETURNING id"
    ), {'bid': bank_id, 'name': display_name}).fetchone()[0]
    conn.commit()

    success = 0
    skipped = 0
    for idx, row in enumerate(rows):
        item_key = row.get('item_key', '')
        if item_key and item_key_exists(item_key):
            skipped += 1
            continue

        qtype = row.get('question_type', 'short_answer')
        qcode = f'{code}_{item_key or f"q{idx:04d}"}'
        stem = row['stem']
        score = Decimal(str(row.get('score', '1') or '1'))

        question_id = conn.execute(text(
            "INSERT INTO qbank_v2_question "
            "(code, owner_id, visibility, origin_type, status, stem, content_format, "
            "question_type, option_data, default_score, created_by, created_time) "
            "VALUES (:code, NULL, 'public', 'imported', 'active', :stem, 'html', "
            ":qtype, CAST(:opts AS jsonb), :score, 1, NOW()) RETURNING id"
        ), {'code': qcode, 'stem': stem, 'qtype': qtype,
            'opts': json.dumps(parse_options_v2(row), ensure_ascii=False), 'score': float(score)}).fetchone()[0]

        conn.execute(text(
            "INSERT INTO qbank_v2_question_answer "
            "(question_id, answer_data, grading_method, grading_config, created_by, created_time) "
            "VALUES (:qid, CAST(:ad AS jsonb), 'exact', '{}'::jsonb, 1, NOW())"
        ), {'qid': question_id, 'ad': json.dumps(parse_answer_v2(row.get('answer', ''), qtype), ensure_ascii=False)})

        is_first = True
        for exp in parse_explanations(row):
            conn.execute(text(
                "INSERT INTO qbank_v2_question_explanation "
                "(question_id, content, explanation_type, is_default, status, created_by, created_time) "
                "VALUES (:qid, :content, :etype, :is_default, 'published', 1, NOW())"
            ), {'qid': question_id, 'content': exp['content'], 'etype': exp['explanation_type'], 'is_default': is_first})
            is_first = False

        # 章节：申论
        sec = conn.execute(text(
            "SELECT id FROM qbank_v2_bank_section "
            "WHERE bank_revision_id = :rid AND code = 'l0/申论' AND deleted = 0"
        ), {'rid': revision_id}).fetchone()
        if sec:
            section_id = sec[0]
        else:
            section_id = conn.execute(text(
                "INSERT INTO qbank_v2_bank_section "
                "(bank_revision_id, code, name, parent_id, depth, sort_order, created_by, created_time) "
                "VALUES (:rid, 'l0/申论', '申论', NULL, 0, 0, 1, NOW()) RETURNING id"
            ), {'rid': revision_id}).fetchone()[0]

        conn.execute(text(
            "INSERT INTO qbank_v2_bank_item "
            "(bank_revision_id, item_key, question_id, section_id, score, sort_order, "
            "is_required, is_active, settings, created_by, created_time) "
            "VALUES (:rid, :ik, :qid, :sid, :score, :sort, true, true, '{}'::jsonb, 1, NOW())"
        ), {'rid': revision_id, 'ik': item_key or f'q{idx:04d}', 'qid': question_id,
            'sid': section_id, 'score': float(score), 'sort': idx})
        success += 1
        conn.commit()

    conn.execute(text("UPDATE qbank_v2_bank_revision SET question_count = :cnt WHERE id = :id"), {'cnt': success, 'id': revision_id})
    conn.commit()
    print(f'    Imported {success} questions, {skipped} skipped')


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--prod-url', required=True, help='生产库连接 URL')
    parser.add_argument('--copy-to', default=None, help='复制 Excel 到的目录')
    args = parser.parse_args()

    prod_engine = get_engine(args.prod_url)
    output_dir = Path(__file__).resolve().parent / 'outputs'
    output_dir.mkdir(exist_ok=True)

    print('Fetching national 申论 banks...')
    with prod_engine.connect() as conn:
        banks = fetch_banks(conn)
    print(f'  Found {len(banks)} banks\n')

    print('Exporting to Excel...')
    files = export_to_excel(prod_engine, banks, output_dir)
    print(f'\nExported {len(files)} files\n')

    print('Importing to local V2...')
    dev_engine = get_engine()
    with dev_engine.connect() as conn:
        fix_sequences(conn)
        for f in files:
            import_file(conn, f)
    print('\nImport done')

    if args.copy_to:
        dst = Path(args.copy_to)
        dst.mkdir(parents=True, exist_ok=True)
        for f in files:
            import shutil
            shutil.copy2(f, dst / f.name)
        print(f'Copied {len(files)} files to {dst}')


if __name__ == '__main__':
    main()
