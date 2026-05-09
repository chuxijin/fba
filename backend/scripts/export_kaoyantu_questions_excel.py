#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
考研兔题目列表导出 Excel 脚本

用法:
    python backend/scripts/export_kaoyantu_questions_excel.py
    python backend/scripts/export_kaoyantu_questions_excel.py --output backend/scripts/outputs/questions.xlsx
    python backend/scripts/export_kaoyantu_questions_excel.py --single-page --page 1
"""
import argparse
import json
import time
from pathlib import Path
from typing import Any

import httpx
from openpyxl import load_workbook
from openpyxl.styles import Alignment

BASE_URL = 'https://api.kaoyantu.top'
PATH = '/routine/auth_api/get_question_list_by_page'
EARMARK_PATH = '/routine/auth_api/get_earmark_list'
DEFAULT_TEMPLATE = r'D:\100_Work\101_Program\Proj\excel\question_import_template.xlsx'
DEFAULT_OUTPUT = 'backend/scripts/outputs/kaoyantu_questions_738_all_pages.xlsx'

DEFAULT_ID = '738'
DEFAULT_PAGE = 1
DEFAULT_LIMIT = 100
DEFAULT_OPEN_ID = 'oRWCb6xMvK5Gj7bmFH_ZNBxB978g'
DEFAULT_TIMESTAMP = 1778134086606
DEFAULT_SIGN = '29B546C59DEBC8BE5940728C1DC32DB2'
DEFAULT_EARMARK_TIMESTAMP = 1778136593570
DEFAULT_EARMARK_SIGN = 'C314D92AB3F9599D8BABC23404387D6F'

DEFAULT_PAGE_PAYLOADS: list[dict[str, str | int]] = [
    {
        'id': '738',
        'page': 1,
        'limit': 100,
        'openId': DEFAULT_OPEN_ID,
        'timeStamp': 1778134086606,
        'sign': '29B546C59DEBC8BE5940728C1DC32DB2',
    },
    {
        'id': '738',
        'page': 2,
        'limit': 100,
        'openId': DEFAULT_OPEN_ID,
        'timeStamp': 1778134086799,
        'sign': '6BE10C938541F392CA9D00F42DA6FDEF',
    },
    {
        'id': '738',
        'page': 3,
        'limit': 100,
        'openId': DEFAULT_OPEN_ID,
        'timeStamp': 1778134086896,
        'sign': '80C9D0E5C4F96483F3C350D1A098F3DA',
    },
    {
        'id': '738',
        'page': 4,
        'limit': 100,
        'openId': DEFAULT_OPEN_ID,
        'timeStamp': 1778134086990,
        'sign': '08094A42CC9C51481D0BD4133FE613F4',
    },
    {
        'id': '738',
        'page': 5,
        'limit': 100,
        'openId': DEFAULT_OPEN_ID,
        'timeStamp': 1778134087079,
        'sign': '59633E6D10171F3C23A5FFEEBF69B019',
    },
    {
        'id': '738',
        'page': 6,
        'limit': 100,
        'openId': DEFAULT_OPEN_ID,
        'timeStamp': 1778134087186,
        'sign': '130D1498649A9AC433EE1AF0923DB4DD',
    },
    {
        'id': '738',
        'page': 7,
        'limit': 100,
        'openId': DEFAULT_OPEN_ID,
        'timeStamp': 1778134087283,
        'sign': '21C456EEB2CF14DFDB9ECCF153F60BBB',
    },
    {
        'id': '738',
        'page': 8,
        'limit': 100,
        'openId': DEFAULT_OPEN_ID,
        'timeStamp': 1778134087376,
        'sign': '22C7B27911711FC221786CAC3A1623C3',
    },
    {
        'id': '738',
        'page': 9,
        'limit': 100,
        'openId': DEFAULT_OPEN_ID,
        'timeStamp': 1778134087486,
        'sign': '695FCDBCB5C6A2DEE8470AACAF4B76C2',
    },
    {
        'id': '738',
        'page': 10,
        'limit': 100,
        'openId': DEFAULT_OPEN_ID,
        'timeStamp': 1778134087579,
        'sign': '737727AD9AC5E55E4DD6050BF05DADC9',
    },
    {
        'id': '738',
        'page': 11,
        'limit': 100,
        'openId': DEFAULT_OPEN_ID,
        'timeStamp': 1778134087675,
        'sign': '99C263ADCFA2846E0A2A6A7469E3F680',
    },
    {
        'id': '738',
        'page': 12,
        'limit': 100,
        'openId': DEFAULT_OPEN_ID,
        'timeStamp': 1778134087778,
        'sign': 'B5976604C6526C3833403A2D14998E6C',
    },
    {
        'id': '738',
        'page': 13,
        'limit': 100,
        'openId': DEFAULT_OPEN_ID,
        'timeStamp': 1778134087877,
        'sign': '64AFF627205C8487E6DFFDAF3ABBE7FE',
    },
]

QUESTION_HEADERS = [
    '序号',
    '题型',
    '题目',
    '选项A',
    '选项B',
    '选项C',
    '选项D',
    '答案',
    '解析',
    '难度',
    '分数',
    '一级目录',
    '二级目录',
    '三级目录',
    '知识点',
    '材料编号',
]


def build_headers() -> dict[str, str]:
    """构建小程序请求头"""
    return {
        'User-Agent': (
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
            '(KHTML, like Gecko) Chrome/132.0.0.0 Safari/537.36 '
            'MicroMessenger/7.0.20.1781(0x6700143B) NetType/WIFI '
            'MiniProgramEnv/Windows WindowsWechat/WMPF WindowsWechat(0x63090a13) '
            'UnifiedPCWindowsWechat(0xf2541843) XWEB/19339'
        ),
        'Xweb_xhr': '1',
        'Content-Type': 'application/json',
        'Accept': '*/*',
        'Sec-Fetch-Site': 'cross-site',
        'Sec-Fetch-Mode': 'cors',
        'Sec-Fetch-Dest': 'empty',
        'Referer': 'https://servicewechat.com/wxaee1a56d07277294/91/page-frame.html',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        'Priority': 'u=1, i',
    }


def build_payload(args: argparse.Namespace) -> dict[str, str | int]:
    """
    构建接口请求体

    :param args: 命令行参数
    :return: 不添加返回说明
    """
    timestamp = args.timestamp
    if args.use_current_timestamp:
        timestamp = int(time.time() * 1000)

    return {
        'id': args.id,
        'page': args.page,
        'limit': args.limit,
        'openId': args.open_id,
        'timeStamp': timestamp,
        'sign': args.sign,
    }


def fetch_question_page(client: httpx.Client, url: str, payload: dict[str, str | int]) -> dict[str, Any]:
    """
    拉取单页题目列表数据

    :param client: HTTP 客户端
    :param url: 接口地址
    :param payload: 请求体
    :return: 不添加返回说明
    """
    response = client.post(url, json=payload)

    response.raise_for_status()
    result = response.json()
    if result.get('code') != 200:
        message = result.get('msg') or '接口返回失败'
        page = payload.get('page')
        raise RuntimeError(f'获取第 {page} 页题目失败: {message}')

    data = result.get('data')
    if not isinstance(data, dict):
        raise RuntimeError('接口 data 结构异常')

    return data


def build_page_payloads(args: argparse.Namespace) -> list[dict[str, str | int]]:
    """
    构建题目分页请求体列表

    :param args: 命令行参数
    :return: 不添加返回说明
    """
    if args.single_page:
        return [build_payload(args)]

    return [
        {
            **payload,
            'id': args.id,
            'openId': args.open_id,
        }
        for payload in DEFAULT_PAGE_PAYLOADS
    ]


def merge_question_pages(page_data_list: list[dict[str, Any]]) -> dict[str, Any]:
    """
    合并分页题目数据

    :param page_data_list: 分页 data 列表
    :return: 不添加返回说明
    """
    if not page_data_list:
        raise RuntimeError('没有可合并的分页数据')

    merged_data = dict(page_data_list[0])
    question_list: list[dict[str, Any]] = []
    question_id_list: list[str] = []
    seen_question_ids: set[int] = set()
    duplicate_count = 0

    for data in page_data_list:
        raw_questions = data.get('question_list')
        if not isinstance(raw_questions, list):
            raise RuntimeError('接口 question_list 结构异常')

        for question in raw_questions:
            if not isinstance(question, dict):
                continue

            question_id = question.get('id')
            if isinstance(question_id, int):
                if question_id in seen_question_ids:
                    duplicate_count += 1
                    continue
                seen_question_ids.add(question_id)
                question_id_list.append(str(question_id))

            question_list.append(question)

    merged_data['question_list'] = question_list
    merged_data['question_id'] = question_id_list
    merged_data['export_duplicate_count'] = duplicate_count
    return merged_data


def fetch_question_data(args: argparse.Namespace) -> dict[str, Any]:
    """
    拉取题目列表数据

    :param args: 命令行参数
    :return: 不添加返回说明
    """
    url = f'{args.base_url.rstrip("/")}{PATH}'
    page_payloads = build_page_payloads(args)
    page_data_list: list[dict[str, Any]] = []

    with httpx.Client(timeout=args.timeout, http2=True, headers=build_headers()) as client:
        for payload in page_payloads:
            page_data_list.append(fetch_question_page(client=client, url=url, payload=payload))

    return merge_question_pages(page_data_list)


def build_earmark_payload(args: argparse.Namespace) -> dict[str, str | int]:
    """
    构建章节树请求体

    :param args: 命令行参数
    :return: 不添加返回说明
    """
    return {
        'cid': args.id,
        'openId': args.open_id,
        'timeStamp': args.earmark_timestamp,
        'sign': args.earmark_sign,
    }


def fetch_earmark_data(args: argparse.Namespace) -> list[dict[str, Any]]:
    """
    拉取章节树数据

    :param args: 命令行参数
    :return: 不添加返回说明
    """
    url = f'{args.base_url.rstrip("/")}{EARMARK_PATH}'
    payload = build_earmark_payload(args)

    with httpx.Client(timeout=args.timeout, http2=True, headers=build_headers()) as client:
        response = client.post(url, json=payload)

    response.raise_for_status()
    result = response.json()
    if result.get('code') != 200:
        message = result.get('msg') or '章节树接口返回失败'
        raise RuntimeError(f'获取章节树失败: {message}')

    data = result.get('data')
    if not isinstance(data, dict):
        raise RuntimeError('章节树 data 结构异常')

    item_list = data.get('list')
    if not isinstance(item_list, list):
        raise RuntimeError('章节树 list 结构异常')

    return [item for item in item_list if isinstance(item, dict)]


def build_chapter_path_map(item_list: list[dict[str, Any]]) -> dict[int, tuple[str, str, str]]:
    """
    构建章节层级映射

    :param item_list: 章节树节点列表
    :return: 不添加返回说明
    """
    item_map: dict[int, dict[str, Any]] = {}
    for item in item_list:
        item_id = item.get('id')
        if isinstance(item_id, int):
            item_map[item_id] = item

    path_map: dict[int, tuple[str, str, str]] = {}
    for item_id, item in item_map.items():
        names: list[str] = []
        current: dict[str, Any] | None = item
        while current:
            name = current.get('name')
            if name:
                names.append(str(name).replace('\t', ' ').strip())

            parent_id = current.get('parent_id')
            if not isinstance(parent_id, int):
                break
            current = item_map.get(parent_id)

        names.reverse()
        while len(names) < 3:
            names.append('')

        path_map[item_id] = (names[0], names[1], names[2])

    return path_map


def normalize_question_type(raw_type: Any) -> str:
    """
    转换题型

    :param raw_type: 接口题型值
    :return: 不添加返回说明
    """
    type_mapping = {
        1: '单选',
        2: '多选',
        3: '判断',
        4: '填空',
        5: '简答',
    }
    if raw_type in type_mapping:
        return type_mapping[raw_type]

    return '单选'


def normalize_difficulty(raw_difficulty: Any) -> str:
    """
    转换难度

    :param raw_difficulty: 接口难度值
    :return: 不添加返回说明
    """
    difficulty_mapping = {
        1: '简单',
        2: '中等',
        3: '困难',
        '1': '简单',
        '2': '中等',
        '3': '困难',
        'easy': '简单',
        'medium': '中等',
        'hard': '困难',
    }
    if raw_difficulty in difficulty_mapping:
        return difficulty_mapping[raw_difficulty]

    return '中等'


def extract_analysis(question: dict[str, Any]) -> str:
    """
    提取解析内容

    :param question: 接口题目数据
    :return: 不添加返回说明
    """
    answer_richtext = question.get('answer_richtext')
    if answer_richtext:
        return str(answer_richtext)

    comments_img = question.get('comments_img')
    if comments_img:
        return str(comments_img)

    return ''


def build_excel_row(
    index: int,
    question: dict[str, Any],
    data: dict[str, Any],
    chapter_path_map: dict[int, tuple[str, str, str]],
) -> list[Any]:
    """
    构建 Excel 行数据

    :param index: 题目序号
    :param question: 接口题目数据
    :param data: 接口 data 数据
    :return: 不添加返回说明
    """
    options = question.get('options') or []
    option_values = list(options[:4])
    while len(option_values) < 4:
        option_values.append('')

    level1_name = data.get('category_name') or data.get('category_short_name') or ''
    level2_name = ''
    level3_name = ''
    question_type = normalize_question_type(question.get('type'))
    eid = question.get('eid')
    if isinstance(eid, int) and eid in chapter_path_map:
        level1_name, level2_name, level3_name = chapter_path_map[eid]
        if level3_name in {'单选', '多选', '判断', '填空', '简答'}:
            question_type = level3_name

    return [
        index,
        question_type,
        question.get('title') or '',
        option_values[0],
        option_values[1],
        option_values[2],
        option_values[3],
        question.get('answer') or '',
        extract_analysis(question),
        normalize_difficulty(question.get('difficulty')),
        1,
        level1_name,
        level2_name,
        level3_name,
        question.get('original_book_number') or '',
        '',
    ]


def clear_sheet_rows(ws: Any, start_row: int = 2) -> None:
    """
    清空模板数据行

    :param ws: 工作表对象
    :param start_row: 起始行
    :return: 不添加返回说明
    """
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            cell.value = None


def write_questions_to_template(
    data: dict[str, Any],
    chapter_path_map: dict[int, tuple[str, str, str]],
    template_path: Path,
    output_path: Path,
) -> int:
    """
    写入题目到模板

    :param data: 接口 data 数据
    :param template_path: 模板路径
    :param output_path: 输出路径
    :return: 不添加返回说明
    """
    workbook = load_workbook(template_path)
    question_sheet = workbook['题目']
    material_sheet = workbook['材料']

    clear_sheet_rows(question_sheet)
    clear_sheet_rows(material_sheet)

    question_list = data.get('question_list')
    if not isinstance(question_list, list):
        raise RuntimeError('接口 question_list 结构异常')

    header_values = [question_sheet.cell(row=1, column=column_index).value for column_index in range(1, 17)]
    if header_values != QUESTION_HEADERS:
        raise RuntimeError(f'模板题目表头不匹配: {header_values}')

    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    for index, question in enumerate(question_list, start=1):
        if not isinstance(question, dict):
            continue

        row_values = build_excel_row(
            index=index,
            question=question,
            data=data,
            chapter_path_map=chapter_path_map,
        )
        row_number = index + 1
        for column_index, value in enumerate(row_values, start=1):
            cell = question_sheet.cell(row=row_number, column=column_index)
            cell.value = value
            cell.alignment = wrap_alignment

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return len(question_list)


def main() -> None:
    """导出题目 Excel"""
    parser = argparse.ArgumentParser(description='导出考研兔题目到导入模板 Excel')
    parser.add_argument('--base-url', default=BASE_URL, help='接口基础地址')
    parser.add_argument('--id', default=DEFAULT_ID, help='题库或列表 ID')
    parser.add_argument('--page', type=int, default=DEFAULT_PAGE, help='页码')
    parser.add_argument('--limit', type=int, default=DEFAULT_LIMIT, help='每页数量')
    parser.add_argument('--open-id', default=DEFAULT_OPEN_ID, help='微信 openId')
    parser.add_argument('--timestamp', type=int, default=DEFAULT_TIMESTAMP, help='请求 timeStamp')
    parser.add_argument('--sign', default=DEFAULT_SIGN, help='请求签名')
    parser.add_argument('--earmark-timestamp', type=int, default=DEFAULT_EARMARK_TIMESTAMP, help='章节树请求 timeStamp')
    parser.add_argument('--earmark-sign', default=DEFAULT_EARMARK_SIGN, help='章节树请求签名')
    parser.add_argument('--single-page', action='store_true', help='只导出 --page 指定的单页')
    parser.add_argument('--use-current-timestamp', action='store_true', help='使用当前毫秒时间戳覆盖 --timestamp')
    parser.add_argument('--timeout', type=float, default=20.0, help='请求超时时间（秒）')
    parser.add_argument('--template', default=DEFAULT_TEMPLATE, help='Excel 模板路径')
    parser.add_argument('--output', default=DEFAULT_OUTPUT, help='输出 Excel 路径')
    args = parser.parse_args()

    data = fetch_question_data(args)
    chapter_path_map = build_chapter_path_map(fetch_earmark_data(args))
    count = write_questions_to_template(
        data=data,
        chapter_path_map=chapter_path_map,
        template_path=Path(args.template),
        output_path=Path(args.output),
    )
    print(json.dumps({
        'output': args.output,
        'count': count,
        'duplicate_count': data.get('export_duplicate_count', 0),
    }, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()
