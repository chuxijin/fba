#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从 Burp 导出文件中提取考研兔题目响应

用法:
    python backend/scripts/extract_kaoyantu_from_burp.py --input burp_items.xml
    python backend/scripts/extract_kaoyantu_from_burp.py --input old.xml patch.xml --output backend/scripts/outputs/merged.xlsx
    python backend/scripts/extract_kaoyantu_from_burp.py --input burp_raw.txt --output backend/scripts/outputs/burp_questions.xlsx
"""
import argparse
import base64
import json
import re
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment

DEFAULT_TEMPLATE = r'D:\100_Work\101_Program\Proj\excel\question_import_template.xlsx'
DEFAULT_OUTPUT = 'backend/scripts/outputs/kaoyantu_burp_questions.xlsx'
DEFAULT_EMPTY_ANALYSIS = '暂无解析'
TARGET_HOST = 'api.kaoyantu.top'
TARGET_PATHS = {
    '/routine/auth_api/get_question_list_by_page',
    '/routine/auth_api/get_earmark_list',
    '/routine/auth_api/get_question_ids',
}

QUESTION_HEADERS = [
    '序号',
    '题型',
    '题目',
    '选项A',
    '选项B',
    '选项C',
    '选项D',
    '答案',
    '解析',
    '难度',
    '分数',
    '一级目录',
    '二级目录',
    '三级目录',
    '知识点',
    '材料编号',
]

QUESTION_TYPE_NAMES = {
    '单选',
    '单选题',
    '多选',
    '多选题',
    '选择题',
    '判断',
    '判断题',
    '填空',
    '填空题',
    '简答',
    '简答题',
    '解答',
    '解答题',
    '分析题',
}

QUESTION_TYPE_MAPPING = {
    '单选': '单选',
    '单选题': '单选',
    '多选': '多选',
    '多选题': '多选',
    '判断': '判断',
    '判断题': '判断',
    '填空': '填空',
    '填空题': '填空',
    '简答': '简答',
    '简答题': '简答',
    '解答': '简答',
    '解答题': '简答',
    '分析题': '简答',
}


def read_text(path: Path) -> str:
    """
    读取文本文件

    :param path: 文件路径
    :return: 不添加返回说明
    """
    raw = path.read_bytes()
    for encoding in ('utf-8', 'utf-8-sig', 'gbk'):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue

    return raw.decode('utf-8', errors='ignore')


def decode_burp_response(response_node: ET.Element) -> str:
    """
    解码 Burp XML 响应节点

    :param response_node: response 节点
    :return: 不添加返回说明
    """
    text = response_node.text or ''
    if response_node.attrib.get('base64') == 'true':
        return base64.b64decode(text).decode('utf-8', errors='ignore')

    return text


def extract_http_body(raw_response: str) -> str:
    """
    提取 HTTP 响应正文

    :param raw_response: 原始 HTTP 响应
    :return: 不添加返回说明
    """
    separators = ('\r\n\r\n', '\n\n')
    for separator in separators:
        if separator in raw_response:
            return raw_response.split(separator, 1)[1].strip()

    return raw_response.strip()


def extract_bodies_from_burp_xml(path: Path) -> tuple[list[str], dict[str, int]]:
    """
    从 Burp XML 提取响应正文

    :param path: XML 文件路径
    :return: 不添加返回说明
    """
    tree = ET.parse(path)
    bodies: list[str] = []
    stats = {
        'burp_items': 0,
        'target_items': 0,
        'skipped_items': 0,
        'empty_response_items': 0,
    }
    for item in tree.findall('.//item'):
        stats['burp_items'] += 1
        host = (item.findtext('host') or '').strip()
        request_path = (item.findtext('path') or '').strip()
        if host != TARGET_HOST or request_path not in TARGET_PATHS:
            stats['skipped_items'] += 1
            continue

        stats['target_items'] += 1
        response_node = item.find('response')
        if response_node is None:
            stats['empty_response_items'] += 1
            continue

        raw_response = decode_burp_response(response_node)
        body = extract_http_body(raw_response)
        if body:
            bodies.append(body)
            continue

        stats['empty_response_items'] += 1

    return bodies, stats


def extract_balanced_json_objects(text: str) -> list[str]:
    """
    从大段文本中提取完整 JSON 对象

    :param text: 原始文本
    :return: 不添加返回说明
    """
    objects: list[str] = []
    start_index: int | None = None
    depth = 0
    in_string = False
    escape = False

    for index, char in enumerate(text):
        if start_index is None:
            if char == '{':
                start_index = index
                depth = 1
            continue

        if escape:
            escape = False
            continue

        if char == '\\':
            escape = True
            continue

        if char == '"':
            in_string = not in_string
            continue

        if in_string:
            continue

        if char == '{':
            depth += 1
            continue

        if char == '}':
            depth -= 1
            if depth == 0:
                objects.append(text[start_index:index + 1])
                start_index = None

    return objects


def extract_bodies_from_raw_text(path: Path) -> tuple[list[str], dict[str, int]]:
    """
    从普通文本提取 JSON 正文

    :param path: 文本路径
    :return: 不添加返回说明
    """
    text = read_text(path)
    bodies: list[str] = []
    for json_text in extract_balanced_json_objects(text):
        try:
            json.loads(json_text)
        except json.JSONDecodeError:
            continue
        bodies.append(json_text)

    return bodies, {
        'burp_items': 0,
        'target_items': len(bodies),
        'skipped_items': 0,
        'empty_response_items': 0,
    }


def load_response_objects(input_path: Path) -> tuple[list[dict[str, Any]], dict[str, int]]:
    """
    加载响应 JSON 对象

    :param input_path: Burp 导出文件
    :return: 不添加返回说明
    """
    head = read_text(input_path)[:200].lstrip()
    if input_path.suffix.lower() == '.xml' or head.startswith('<?xml') or head.startswith('<items'):
        bodies, stats = extract_bodies_from_burp_xml(input_path)
    else:
        bodies, stats = extract_bodies_from_raw_text(input_path)

    responses: list[dict[str, Any]] = []
    json_error_count = 0
    for body in bodies:
        try:
            parsed = json.loads(body)
        except json.JSONDecodeError:
            json_error_count += 1
            continue

        if isinstance(parsed, dict):
            responses.append(parsed)

    stats['json_error_count'] = json_error_count
    return responses, stats


def load_all_response_objects(input_paths: list[Path]) -> tuple[list[dict[str, Any]], dict[str, int]]:
    """
    加载多个输入文件响应

    :param input_paths: 输入文件路径列表
    :return: 不添加返回说明
    """
    all_responses: list[dict[str, Any]] = []
    merged_stats = {
        'input_files': len(input_paths),
        'burp_items': 0,
        'target_items': 0,
        'skipped_items': 0,
        'empty_response_items': 0,
        'json_error_count': 0,
    }
    for input_path in input_paths:
        responses, stats = load_response_objects(input_path)
        all_responses.extend(responses)
        for key in ('burp_items', 'target_items', 'skipped_items', 'empty_response_items', 'json_error_count'):
            merged_stats[key] += stats.get(key, 0)

    return all_responses, merged_stats


def normalize_question_type(question: dict[str, Any], chapter_names: list[str] | None = None) -> str:
    """
    转换题型

    :param question: 题目数据
    :param chapter_names: 章节名称列表
    :return: 不添加返回说明
    """
    answer = str(question.get('answer') or '')
    if len(re.findall(r'[A-Z]', answer.upper())) > 1:
        return '多选'

    raw_type = question.get('type')
    type_mapping = {
        1: '单选',
        2: '多选',
        3: '判断',
        4: '填空',
        5: '简答',
    }
    if raw_type in type_mapping:
        return type_mapping[raw_type]

    if chapter_names:
        for chapter_name in reversed(chapter_names):
            if chapter_name in QUESTION_TYPE_MAPPING:
                return QUESTION_TYPE_MAPPING[chapter_name]

    return '单选'


def has_multiple_option_answer(question: dict[str, Any]) -> bool:
    """
    判断是否为多选答案

    :param question: 题目数据
    :return: 不添加返回说明
    """
    answer = str(question.get('answer') or '')
    return len(re.findall(r'[A-Z]', answer.upper())) > 1


def clean_chapter_names(chapter_names: list[str]) -> list[str]:
    """
    清洗章节目录

    :param chapter_names: 原始章节名称列表
    :return: 不添加返回说明
    """
    return [name for name in chapter_names if name and name not in QUESTION_TYPE_NAMES]


def build_chapter_tuple(chapter_names: list[str]) -> tuple[str, str, str, str, str]:
    """
    构建章节元组

    :param chapter_names: 原始章节名称列表
    :return: 不添加返回说明
    """
    clean_names = clean_chapter_names(chapter_names)
    knowledge_name = ''
    if clean_names:
        knowledge_name = clean_names[0]
    elif chapter_names:
        knowledge_name = chapter_names[0]

    padded_names = clean_names[:3]
    while len(padded_names) < 3:
        padded_names.append('')

    question_type = normalize_question_type({}, chapter_names)
    return padded_names[0], padded_names[1], padded_names[2], knowledge_name, question_type


def normalize_difficulty(raw_difficulty: Any) -> str:
    """
    转换难度

    :param raw_difficulty: 接口难度值
    :return: 不添加返回说明
    """
    difficulty_mapping = {
        1: '简单',
        2: '中等',
        3: '困难',
        '1': '简单',
        '2': '中等',
        '3': '困难',
        'easy': '简单',
        'medium': '中等',
        'hard': '困难',
    }
    if raw_difficulty in difficulty_mapping:
        return difficulty_mapping[raw_difficulty]

    return '中等'


def normalize_score(question_type: str) -> int:
    """
    转换分数

    :param question_type: 题型
    :return: 不添加返回说明
    """
    if question_type == '多选':
        return 2

    return 1


def extract_analysis(
    question: dict[str, Any],
    analysis_map: dict[int, str] | None = None,
    *,
    use_fallback: bool = True,
) -> str:
    """
    提取解析内容

    :param question: 题目数据
    :param analysis_map: 题目 ID 与解析内容映射
    :param use_fallback: 是否使用默认空解析文案
    :return: 不添加返回说明
    """
    question_id = question.get('id')
    if isinstance(question_id, int) and analysis_map and question_id in analysis_map:
        return analysis_map[question_id]

    candidates = [
        question.get('analysis'),
        question.get('解析'),
        question.get('explain'),
        question.get('explanation'),
        question.get('comment'),
        question.get('comments'),
        question.get('answer_richtext'),
        question.get('comments_img'),
    ]
    for candidate in candidates:
        if candidate:
            return str(candidate)

    if use_fallback:
        return DEFAULT_EMPTY_ANALYSIS

    return ''


def collect_analysis_map(responses: list[dict[str, Any]]) -> dict[int, str]:
    """
    收集题目解析映射

    :param responses: 响应对象列表
    :return: 不添加返回说明
    """
    analysis_map: dict[int, str] = {}
    for response in responses:
        data = response.get('data')
        if not isinstance(data, list):
            continue

        for item in data:
            if not isinstance(item, dict):
                continue

            question_id = item.get('id')
            comments = item.get('comments')
            if isinstance(question_id, int) and comments:
                analysis_map[question_id] = str(comments)

    return analysis_map


def build_chapter_path_map(responses: list[dict[str, Any]]) -> dict[int, tuple[str, str, str, str, str]]:
    """
    从响应中构建章节映射

    :param responses: 响应对象列表
    :return: 不添加返回说明
    """
    item_map: dict[int, dict[str, Any]] = {}
    for response in responses:
        data = response.get('data')
        if not isinstance(data, dict):
            continue

        item_list = data.get('list')
        if not isinstance(item_list, list):
            continue

        for item in item_list:
            if not isinstance(item, dict):
                continue

            item_id = item.get('id')
            if isinstance(item_id, int) and 'parent_id' in item and 'name' in item:
                item_map[item_id] = item

    path_map: dict[int, tuple[str, str, str, str, str]] = {}
    for item_id, item in item_map.items():
        names: list[str] = []
        current: dict[str, Any] | None = item
        while current:
            name = current.get('name')
            if name:
                names.append(str(name).replace('\t', ' ').strip())

            parent_id = current.get('parent_id')
            if not isinstance(parent_id, int):
                break
            current = item_map.get(parent_id)

        names.reverse()
        path_map[item_id] = build_chapter_tuple(names)

    return path_map


def find_question_dicts(value: Any) -> list[dict[str, Any]]:
    """
    递归查找题目对象

    :param value: 任意响应结构
    :return: 不添加返回说明
    """
    questions: list[dict[str, Any]] = []
    if isinstance(value, dict):
        if value.get('title') and value.get('options') and value.get('answer'):
            questions.append(value)

        for child in value.values():
            questions.extend(find_question_dicts(child))
        return questions

    if isinstance(value, list):
        for item in value:
            questions.extend(find_question_dicts(item))

    return questions


def collect_questions(responses: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """
    收集并去重题目

    :param responses: 响应对象列表
    :return: 不添加返回说明
    """
    seen_ids: set[int] = set()
    questions: list[dict[str, Any]] = []
    for response in responses:
        for question in find_question_dicts(response):
            question_id = question.get('id')
            if isinstance(question_id, int):
                if question_id in seen_ids:
                    continue
                seen_ids.add(question_id)

            questions.append(question)

    return questions


def build_missing_analysis_ids(questions: list[dict[str, Any]], analysis_map: dict[int, str]) -> list[int]:
    """
    构建缺失解析题目 ID 列表

    :param questions: 题目列表
    :param analysis_map: 解析映射
    :return: 不添加返回说明
    """
    missing_ids: list[int] = []
    for question in questions:
        question_id = question.get('id')
        if not isinstance(question_id, int):
            continue

        if question_id not in analysis_map and not extract_analysis(question, use_fallback=False):
            missing_ids.append(question_id)

    return missing_ids


def filter_questions_with_analysis(
    questions: list[dict[str, Any]],
    analysis_map: dict[int, str],
) -> list[dict[str, Any]]:
    """
    过滤有真实解析的题目

    :param questions: 题目列表
    :param analysis_map: 解析映射
    :return: 不添加返回说明
    """
    filtered_questions: list[dict[str, Any]] = []
    for question in questions:
        question_id = question.get('id')
        if isinstance(question_id, int) and question_id in analysis_map:
            filtered_questions.append(question)
            continue

        if extract_analysis(question, use_fallback=False):
            filtered_questions.append(question)

    return filtered_questions


def clear_sheet_rows(ws: Any, start_row: int = 2) -> None:
    """
    清空模板数据行

    :param ws: 工作表对象
    :param start_row: 起始行
    :return: 不添加返回说明
    """
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            cell.value = None


def build_excel_row(
    index: int,
    question: dict[str, Any],
    chapter_path_map: dict[int, tuple[str, str, str, str, str]],
    analysis_map: dict[int, str],
) -> list[Any]:
    """
    构建 Excel 行

    :param index: 题目序号
    :param question: 题目数据
    :param chapter_path_map: 章节映射
    :param analysis_map: 题目 ID 与解析内容映射
    :return: 不添加返回说明
    """
    options = question.get('options') or []
    option_values = list(options[:4])
    while len(option_values) < 4:
        option_values.append('')

    level1_name = ''
    level2_name = ''
    level3_name = ''
    knowledge_name = ''
    mapped_question_type = ''
    eid = question.get('eid')
    if isinstance(eid, int) and eid in chapter_path_map:
        level1_name, level2_name, level3_name, knowledge_name, mapped_question_type = chapter_path_map[eid]

    question_type = normalize_question_type(question)
    if mapped_question_type and not has_multiple_option_answer(question):
        question_type = mapped_question_type

    return [
        index,
        question_type,
        question.get('title') or '',
        option_values[0],
        option_values[1],
        option_values[2],
        option_values[3],
        question.get('answer') or '',
        extract_analysis(question, analysis_map),
        normalize_difficulty(question.get('difficulty')),
        normalize_score(question_type),
        level1_name,
        level2_name,
        level3_name,
        knowledge_name or question.get('original_book_number') or '',
        '',
    ]


def write_excel(
    questions: list[dict[str, Any]],
    chapter_path_map: dict[int, tuple[str, str, str, str, str]],
    analysis_map: dict[int, str],
    template_path: Path,
    output_path: Path,
) -> None:
    """
    写入 Excel

    :param questions: 题目列表
    :param chapter_path_map: 章节映射
    :param analysis_map: 题目 ID 与解析内容映射
    :param template_path: 模板路径
    :param output_path: 输出路径
    :return: 不添加返回说明
    """
    workbook = load_workbook(template_path)
    question_sheet = workbook['题目']
    material_sheet = workbook['材料']
    clear_sheet_rows(question_sheet)
    clear_sheet_rows(material_sheet)

    header_values = [question_sheet.cell(row=1, column=column_index).value for column_index in range(1, 17)]
    if header_values != QUESTION_HEADERS:
        raise RuntimeError(f'模板题目表头不匹配: {header_values}')

    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    for index, question in enumerate(questions, start=1):
        row_values = build_excel_row(
            index=index,
            question=question,
            chapter_path_map=chapter_path_map,
            analysis_map=analysis_map,
        )
        for column_index, value in enumerate(row_values, start=1):
            cell = question_sheet.cell(row=index + 1, column=column_index)
            cell.value = value
            cell.alignment = wrap_alignment

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> None:
    """执行 Burp 响应提取"""
    parser = argparse.ArgumentParser(description='从 Burp 导出文件提取考研兔题目到 Excel')
    parser.add_argument('--input', required=True, nargs='+', help='Burp 导出的 XML 或普通文本文件，可传多个')
    parser.add_argument('--template', default=DEFAULT_TEMPLATE, help='Excel 模板路径')
    parser.add_argument('--output', default=DEFAULT_OUTPUT, help='输出 Excel 路径')
    parser.add_argument('--only-with-analysis', action='store_true', help='只导出有真实解析的题目，适合补充包')
    args = parser.parse_args()

    input_paths = [Path(input_path) for input_path in args.input]
    responses, load_stats = load_all_response_objects(input_paths)
    chapter_path_map = build_chapter_path_map(responses)
    analysis_map = collect_analysis_map(responses)
    questions = collect_questions(responses)
    all_questions_count = len(questions)
    if args.only_with_analysis:
        questions = filter_questions_with_analysis(questions, analysis_map)

    if not questions:
        raise RuntimeError('未从 Burp 文件中提取到题目数据')

    write_excel(
        questions=questions,
        chapter_path_map=chapter_path_map,
        analysis_map=analysis_map,
        template_path=Path(args.template),
        output_path=Path(args.output),
    )

    analysis_count = sum(1 for question in questions if extract_analysis(question, analysis_map))
    missing_analysis_ids = build_missing_analysis_ids(questions, analysis_map)
    print(json.dumps({
        **load_stats,
        'responses': len(responses),
        'all_questions': all_questions_count,
        'questions': len(questions),
        'filtered_without_analysis': all_questions_count - len(questions),
        'chapter_mappings': len(chapter_path_map),
        'analysis_mappings': len(analysis_map),
        'analysis_count': analysis_count,
        'missing_analysis_count': len(missing_analysis_ids),
        'missing_analysis_ids': missing_analysis_ids[:50],
        'output': args.output,
    }, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()
