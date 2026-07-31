#!/usr/bin/env python3
"""Export 国家公务员考试行测题库 to V2-format XLSX, grouped by chapter.

用法:
  uv run python backend/scripts/export_gwy_xc_to_excel.py
  uv run python backend/scripts/export_gwy_xc_to_excel.py --db-url postgresql+psycopg://user:pass@host:8100/fba
"""

import argparse
import json
import os
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

os.environ['FBA_DEV'] = '1'

from sqlalchemy import create_engine, text
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

COLUMNS = [
    'question_type', 'stem', 'answer', 'explanation_default',
    'option_A', 'option_B', 'option_C', 'option_D', 'option_E',
    'score',
    'section_l1', 'section_l2', 'section_l3',
    'knowledge_point', 'item_key',
]

V1_TO_V2_TYPE = {
    'single': 'single_choice',
    'multiple': 'multiple_choice',
    'judgement': 'true_false',
    'fill': 'fill_blank',
    'shortAnswer': 'short_answer',
}


def get_engine(db_url=None):
    if db_url:
        return create_engine(db_url, pool_pre_ping=True)
    from backend.core.conf import settings
    driver = 'psycopg' if settings.DATABASE_TYPE == 'postgresql' else 'pymysql'
    database = settings.DATABASE_SCHEMA
    url = (
        f"{settings.DATABASE_TYPE}+{driver}://{settings.DATABASE_USER}:{settings.DATABASE_PASSWORD}"
        f"@{settings.DATABASE_HOST}:{settings.DATABASE_PORT}/{database}"
    )
    return create_engine(url, pool_pre_ping=True)


def fetch_chapters(conn):
    rows = conn.execute(text(
        "SELECT id, name, parent_id, level, sort_order "
        "FROM fba.study_question_chapter "
        "WHERE bank_id = 60 AND deleted = 0 "
        "ORDER BY level, sort_order"
    )).mappings().all()
    return [dict(r) for r in rows]


def fetch_placements(conn, chapter_ids):
    if not chapter_ids:
        return []
    rows = conn.execute(text(
        "SELECT p.id, p.question_id, p.chapter_id, p.sort_order, p.bank_id "
        "FROM fba.study_question_placement p "
        "JOIN fba.study_question_bank b ON p.bank_id = b.id "
        "WHERE b.chapter_source_bank_id = 60 "
        "AND p.deleted = 0 AND p.is_active = true AND b.deleted = 0 "
        "AND p.chapter_id = ANY(:cids) "
        "ORDER BY p.chapter_id, p.sort_order"
    ), {'cids': chapter_ids}).mappings().all()
    return [dict(r) for r in rows]


def fetch_questions(conn, question_ids):
    if not question_ids:
        return {}
    rows = conn.execute(text(
        "SELECT q.id, q.type, q.stem, q.options, q.difficulty, q.default_score, q.knowledge_point "
        "FROM fba.study_question q "
        "WHERE q.id = ANY(:qids) AND q.deleted = 0"
    ), {'qids': question_ids}).mappings().all()
    return {r['id']: dict(r) for r in rows}


def fetch_analyses(conn, question_ids):
    if not question_ids:
        return {}
    rows = conn.execute(text(
        "SELECT a.question_id, a.answer_data, a.content "
        "FROM fba.study_question_analysis a "
        "WHERE a.question_id = ANY(:qids) AND a.deleted = 0 AND a.is_default = true"
    ), {'qids': question_ids}).mappings().all()
    return {r['question_id']: dict(r) for r in rows}


def parse_options(options_json):
    if not options_json:
        return {}
    if isinstance(options_json, str):
        opts = json.loads(options_json)
    else:
        opts = options_json
    result = {}
    if isinstance(opts, dict):
        for k, v in opts.items():
            label = k.strip().upper()
            if label and v:
                result[f'option_{label}'] = safe_str(v)
    elif isinstance(opts, list):
        labels = ['A', 'B', 'C', 'D', 'E']
        for i, item in enumerate(opts):
            if i < len(labels) and item:
                if isinstance(item, dict):
                    content = item.get('content') or item.get('label') or str(item)
                else:
                    content = str(item)
                result[f'option_{labels[i]}'] = safe_str(content)
    return result


def parse_answer(answer_data):
    if not answer_data:
        return ''
    if isinstance(answer_data, str):
        data = json.loads(answer_data)
    else:
        data = answer_data
    if isinstance(data, dict):
        answers = data.get('answer') or data.get('answers') or data.get('correct') or ''
        if isinstance(answers, list):
            return safe_str(','.join(str(a) for a in answers))
        return safe_str(answers)
    if isinstance(data, list):
        return safe_str(','.join(str(a) for a in data))
    return safe_str(data)


def get_section_names(chapter_id, chapter_map):
    ch = chapter_map.get(chapter_id)
    if not ch:
        return ['', '', '']
    if ch['level'] == 1:
        return [ch['name'], '', '']
    parent = chapter_map.get(ch['parent_id'])
    parent_name = parent['name'] if parent else ''
    if ch['level'] == 2:
        return [parent_name, ch['name'], '']
    if ch['level'] == 3:
        grandparent = chapter_map.get(parent['parent_id']) if parent else None
        gp_name = grandparent['name'] if grandparent else ''
        return [gp_name, parent_name, ch['name']]
    return ['', '', '']


def safe_str(val):
    """Sanitize string for openpyxl cell (remove illegal chars, truncate)."""
    if val is None:
        return ''
    s = str(val)
    # Remove control chars except \t, \n, \r
    s = ''.join(c if c == '\t' or c == '\n' or c == '\r' or (ord(c) >= 32) else ' ' for c in s)
    # Truncate to 32767 chars (openpyxl limit)
    if len(s) > 32767:
        s = s[:32767]
    return s


def main():
    parser = argparse.ArgumentParser(description='Export 国考行测题库 to XLSX')
    parser.add_argument('--db-url', help='数据库连接 URL (默认从 settings 读取)')
    parser.add_argument('--output', default=None, help='输出文件路径')
    args = parser.parse_args()

    t0 = time.time()
    engine = get_engine(args.db_url)

    with engine.connect() as conn:
        print('Fetching chapters...')
        chapters = fetch_chapters(conn)
        chapter_map = {c['id']: c for c in chapters}
        print(f'  Found {len(chapters)} chapters')

        print('Fetching placements...')
        placements = fetch_placements(conn, [c['id'] for c in chapters])
        print(f'  Found {len(placements)} placements')

        qids = list(set(p['question_id'] for p in placements))
        print(f'Fetching {len(qids)} questions...')
        questions = fetch_questions(conn, qids)
        print(f'  Got {len(questions)} questions')

        print('Fetching analyses...')
        analyses = fetch_analyses(conn, qids)
        print(f'  Got {len(analyses)} analyses')

    wb = Workbook()
    ws = wb.active
    ws.title = '题目'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    row_num = 2
    written = 0

    for ch in chapters:
        ch_placements = [p for p in placements if p['chapter_id'] == ch['id']]
        if not ch_placements:
            print(f'  Chapter "{ch["name"]}" — 0 questions, skipping')
            continue
        print(f'  Chapter "{ch["name"]}" — {len(ch_placements)} questions')

        for p in ch_placements:
            q = questions.get(p['question_id'])
            if not q:
                continue

            v2_type = V1_TO_V2_TYPE.get(q['type'], q['type'])
            options = parse_options(q['options'])
            analysis = analyses.get(p['question_id'], {})
            answer = parse_answer(analysis.get('answer_data'))
            kp = q['knowledge_point']
            kp_str = ''
            if kp:
                if isinstance(kp, str):
                    kp_str = kp
                elif isinstance(kp, list):
                    kp_str = '; '.join(str(x.get('name', '')) for x in kp if isinstance(x, dict))
            sections = get_section_names(p['chapter_id'], chapter_map)

            row_data = {
                'question_type': v2_type,
                'stem': safe_str(q['stem'] or ''),
                'answer': answer,
                'explanation_default': safe_str(analysis.get('content', '')),
                **options,
                'score': safe_str(q.get('default_score', '')),
                'section_l1': sections[0],
                'section_l2': sections[1],
                'section_l3': sections[2],
                'knowledge_point': safe_str(kp_str),
                'item_key': f'gwy_xc_{p["bank_id"]}_{p["question_id"]}',
            }

            for col_idx, col_name in enumerate(COLUMNS, 1):
                val = safe_str(row_data.get(col_name, ''))
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.border = thin_border
                if col_name == 'stem':
                    cell.alignment = Alignment(wrap_text=True)

            row_num += 1
            written += 1

    if args.output:
        output_path = Path(args.output)
    else:
        output_dir = Path(__file__).resolve().parent / 'outputs'
        output_dir.mkdir(exist_ok=True)
        output_path = output_dir / '国考行测题库.xlsx'

    wb.save(output_path)
    elapsed = time.time() - t0
    print(f'\nDone! Written {written} rows to {output_path} ({elapsed:.1f}s)')


if __name__ == '__main__':
    main()
