#!/usr/bin/env python3
"""Export 国考行测各年份试卷到独立 XLSX 文件，按章节分组。"""

import argparse
import json
import os
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
os.environ['FBA_DEV'] = '1'

from sqlalchemy import create_engine, text
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

COLUMNS = [
    'question_type', 'stem', 'answer', 'explanation_default',
    'option_A', 'option_B', 'option_C', 'option_D', 'option_E',
    'score',
    'section_l1', 'section_l2', 'section_l3',
    'knowledge_point', 'item_key',
]

V1_TO_V2_TYPE = {
    'single': 'single_choice', 'multiple': 'multiple_choice',
    'judgement': 'true_false', 'fill': 'fill_blank', 'shortAnswer': 'short_answer',
}


def safe_str(val):
    if val is None:
        return ''
    s = str(val)
    s = ''.join(c if c == '\t' or c == '\n' or c == '\r' or (ord(c) >= 32) else ' ' for c in s)
    if len(s) > 32767:
        s = s[:32767]
    return s


def get_engine(db_url=None):
    if db_url:
        return create_engine(db_url, pool_pre_ping=True)
    from backend.core.conf import settings
    driver = 'psycopg' if settings.DATABASE_TYPE == 'postgresql' else 'pymysql'
    database = settings.DATABASE_SCHEMA
    url = (
        f"{settings.DATABASE_TYPE}+{driver}://{settings.DATABASE_USER}:{settings.DATABASE_PASSWORD}"
        f"@{settings.DATABASE_HOST}:{settings.DATABASE_PORT}/{database}"
    )
    return create_engine(url, pool_pre_ping=True)


def get_national_paper_banks(conn):
    """Get national exam paper banks (PAPER_GWY_XC_NAT_*)."""
    rows = conn.execute(text(
        "SELECT id, code, name, year "
        "FROM fba.study_question_bank "
        "WHERE code LIKE 'PAPER_GWY_XC_NAT_%' AND deleted = 0 AND status = 1 "
        "ORDER BY year DESC, id"
    )).mappings().all()
    return [dict(r) for r in rows]


def get_chapters(conn, bank_id=60):
    rows = conn.execute(text(
        "SELECT id, name, parent_id, level, sort_order "
        "FROM fba.study_question_chapter "
        "WHERE bank_id = :bid AND deleted = 0 "
        "ORDER BY level, sort_order"
    ), {'bid': bank_id}).mappings().all()
    return [dict(r) for r in rows]


def get_placements(conn, bank_id):
    rows = conn.execute(text(
        "SELECT p.question_id, p.chapter_id, p.sort_order "
        "FROM fba.study_question_placement p "
        "WHERE p.bank_id = :bid AND p.deleted = 0 AND p.is_active = true "
        "ORDER BY p.chapter_id, p.sort_order"
    ), {'bid': bank_id}).mappings().all()
    return [dict(r) for r in rows]


def batch_fetch_questions(conn, qids):
    result = {}
    if not qids:
        return result
    BATCH = 5000
    for i in range(0, len(qids), BATCH):
        batch = qids[i:i + BATCH]
        rows = conn.execute(text(
            "SELECT q.id, q.type, q.stem, q.options, q.difficulty, q.default_score, q.knowledge_point "
            "FROM fba.study_question q "
            "WHERE q.id = ANY(:qids) AND q.deleted = 0"
        ), {'qids': batch}).mappings().all()
        for r in rows:
            result[r['id']] = dict(r)
    return result


def batch_fetch_analyses(conn, qids):
    result = {}
    if not qids:
        return result
    BATCH = 5000
    for i in range(0, len(qids), BATCH):
        batch = qids[i:i + BATCH]
        rows = conn.execute(text(
            "SELECT a.question_id, a.answer_data, a.content "
            "FROM fba.study_question_analysis a "
            "WHERE a.question_id = ANY(:qids) AND a.deleted = 0 AND a.is_default = true"
        ), {'qids': batch}).mappings().all()
        for r in rows:
            result[r['question_id']] = dict(r)
    return result


def parse_options(options_json):
    if not options_json:
        return {}
    opts = json.loads(options_json) if isinstance(options_json, str) else options_json
    result = {}
    if isinstance(opts, dict):
        for k, v in opts.items():
            label = k.strip().upper()
            if label and v:
                result[f'option_{label}'] = safe_str(v)
    elif isinstance(opts, list):
        labels = ['A', 'B', 'C', 'D', 'E']
        for i, item in enumerate(opts):
            if i < len(labels) and item:
                content = item.get('content') or item.get('label') or str(item) if isinstance(item, dict) else str(item)
                result[f'option_{labels[i]}'] = safe_str(content)
    return result


def parse_answer(answer_data):
    if not answer_data:
        return ''
    data = json.loads(answer_data) if isinstance(answer_data, str) else answer_data
    if isinstance(data, dict):
        answers = data.get('answer') or data.get('answers') or data.get('correct') or ''
        return safe_str(','.join(str(a) for a in answers) if isinstance(answers, list) else answers)
    return safe_str(','.join(str(a) for a in data) if isinstance(data, list) else str(data))


def sanitize_filename(name):
    s = name.replace('/', '／').replace('\\', '＼').replace(':', '：').replace('*', '＊').replace('?', '？')
    s = s.replace('"', "'").replace('<', '＜').replace('>', '＞').replace('|', '｜')
    return s.strip()


def export_bank(conn, bank, chapters, output_dir):
    t0 = time.time()
    chapter_map = {c['id']: c for c in chapters}

    placements = get_placements(conn, bank['id'])
    if not placements:
        print(f'  [{bank["name"]}] 0 questions, skipped')
        return

    qids = list(set(p['question_id'] for p in placements))
    questions = batch_fetch_questions(conn, qids)
    analyses = batch_fetch_analyses(conn, qids)

    wb = Workbook()
    ws = wb.active
    ws.title = '题目'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    for ci, cn in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=cn)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    row_num = 2
    written = 0

    for ch in chapters:
        ch_placements = [p for p in placements if p['chapter_id'] == ch['id']]
        if not ch_placements:
            continue

        for p in ch_placements:
            q = questions.get(p['question_id'])
            if not q:
                continue

            v2_type = V1_TO_V2_TYPE.get(q['type'], q['type'])
            options = parse_options(q['options'])
            analysis = analyses.get(p['question_id'], {})
            answer = parse_answer(analysis.get('answer_data'))
            kp = q['knowledge_point']
            kp_str = ''
            if kp:
                kp_str = kp if isinstance(kp, str) else '; '.join(
                    str(x.get('name', '')) for x in (kp if isinstance(kp, list) else [kp]) if isinstance(x, dict)
                )
            ch_name = chapter_map.get(p['chapter_id'], {}).get('name', '')

            row_data = {
                'question_type': v2_type,
                'stem': safe_str(q['stem'] or ''),
                'answer': answer,
                'explanation_default': safe_str(analysis.get('content', '')),
                **options,
                'score': safe_str(q.get('default_score', '')),
                'section_l1': ch_name,
                'section_l2': '',
                'section_l3': '',
                'knowledge_point': safe_str(kp_str),
                'item_key': f'nat_{bank["year"]}_{bank["id"]}_{p["question_id"]}',
            }

            for ci, cn in enumerate(COLUMNS, 1):
                val = row_data.get(cn, '')
                cell = ws.cell(row=row_num, column=ci, value=val)
                cell.border = thin_border
                if cn == 'stem':
                    cell.alignment = Alignment(wrap_text=True)

            row_num += 1
            written += 1

    safe_name = sanitize_filename(bank['name'])
    fname = f'{bank["year"]}_{bank["code"]}_{safe_name}.xlsx'
    fpath = output_dir / fname
    wb.save(fpath)
    elapsed = time.time() - t0
    print(f'  [{bank["name"]}] {written} questions → {fname} ({elapsed:.1f}s)')


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--db-url', help='数据库连接 URL')
    parser.add_argument('--output-dir', default=None, help='输出目录')
    args = parser.parse_args()

    output_dir = Path(args.output_dir) if args.output_dir else (
        Path(__file__).resolve().parent / 'outputs'
    )
    output_dir.mkdir(exist_ok=True)

    engine = get_engine(args.db_url)
    with engine.connect() as conn:
        banks = get_national_paper_banks(conn)
        chapters = get_chapters(conn)
        print(f'Found {len(banks)} national exam paper banks, {len(chapters)} chapters\n')

        for bank in banks:
            export_bank(conn, bank, chapters, output_dir)

    print(f'\nAll done! Files in {output_dir}')


if __name__ == '__main__':
    main()
